import os
import io
import sys
import openpyxl
import shutil
import fileinput
import subprocess
import time
import datetime
import argparse
import re
import glob

param_dict = {}
keys_values = {
        'GOP': ['Gop.DoubleRef', 'Gop.EnableLT', 'Gop.FreqIDR', 'Gop.FreqLT', 'Gop.FreqRP', 'Gop.GdrMode', 'Gop.Length',
        'Gop.NumB', 'Gop.TempDQP', 'Gop.WriteAVCHeaderSVC', 'GopCtrlMode'],

        'INPUT': ['CmdFile', 'I|CropHeight', 'I|CropPosX', 'I|CropPosY', 'I|CropWidth', 'I|Format', 'I|FrameRate', 'GMVFile', 'HDRFile',
        'I|Height', 'MapFile', 'QpTablesFolder', 'ROIFile', 'I|Width', 'I|YUVFile'],

        'DYNAMIC_INPUT': ['D|Height', 'D|Width', 'D|YUVFile'],

        'OUTPUT': ['BitstreamFile', 'O|CropHeight', 'O|CropPosX', 'O|CropPosY', 'O|CropWidth', 'O|Format', 'RecFile'],

        'RATE_CONTROL': ['BitRate', 'CPBSize', 'EnableSkip', 'FrameRate', 'InitialDelay', 'IPDelta', 'MaxBitRate',
        'MaxConsecutiveSkip', 'MaxPictureSize', 'MaxPictureSize.B', 'MaxPictureSize.I', 'MaxPictureSize.P',
        'MaxPictureSizeInBits', 'MaxPictureSizeInBits.B', 'MaxPictureSizeInBits.I', 'MaxPictureSizeInBits.P', 'MaxPSNR',
        'MaxQP', 'MaxQP.B', 'MaxQP.I', 'MaxQP.P', 'MinPSNR', 'MinQP', 'MinQP.B', 'MinQP.I',
        'MinQP.P', 'PBDelta', 'RateCtrlMode', 'ScnChgResilience', 'SCPrevention', 'SliceQP', 'UseGoldenRef'],

        'SETTINGS': ['AspectRatio', 'AvcLowLat', 'BitDepth', 'CabacInit', 'ChromaMode', 'ColourDescription', 'ColourMatrix',
        'Compression', 'ConstrainedIntraPred', 'CostMode', 'CuQpDeltaDepth',
        'DependentSlice', 'DirectMode', 'EnableAUD', 'EnableFillerData', 'EnableSEI', 'EntropyMode', 'FileScalingList', 'LambdaCtrlMode',
        'LambdaFactors', 'Level', 'LookAhead', 'LoopFilter', 'LoopFilter.BetaOffset', 'LoopFilter.CrossSlice', 'LoopFilter.CrossTile',
        'LoopFilter.TcOffset', 'NumCore', 'NumSlices', 'PCM', 'PicCbQpOffset', 'PicCrQpOffset', 'Profile', 'QPCtrlMode',
        'SAO', 'ScalingList', 'SCDFirstPass', 'SliceCbQpOffset', 'SliceCrQpOffset', 'SliceLat', 'SrcFormat', 'StartCode',
        'SubframeLatency', 'Tier', 'TransferCharac', 'TwoPass', 'TwoPassFile', 'UniformSliceType', 'UseL2C', 'VideoFullRange',
        'VideoMode', 'WeightedPred'],

        'RUN': ['BitrateFile', 'FirstPicture',
        'InputSleep', 'Loop', 'MaxPicture', 'ScnChgLookAhead', 'UseBoard']
}

for key, values in keys_values.items():
    param_dict[key] = values

error_dict = {}

error_msg = {
        'Error in ctrlsw app': ['Error', 'error', 'ERROR'],
        'Assertion \'0\' failed': ['Assertion|failed', 'Assertion', 'assertion'],
        'Encoder Resource Error': ['Failed to create Encoder'],
        'CMA allocation Error': ['Cannot allocate memory'],
        'No QP File found': ['No QP File found'],
        'Unknown identifire please check property name': ['unknown identifier'],
        'Exception caught: No frame decoded': ['Exception caught: No frame decoded'],
        'I/p YUVFile not found': ['Exception caught: Can\'t open file for reading'],
        'Exception caught Error': ['Exception caught'],
        'Get higher Profile to support the usecase': ['getHevcMinimumProfile: Assertion \`0\' failed']
}

for key, values in error_msg.items():
    error_dict[key] = values


#param_dict[key] = values

def open_workbook(xls_file, xls_sheet):

    #open workbook
    workbook = openpyxl.load_workbook(xls_file)

    #Select the sheet by Index or name
    workbook_sheet = workbook[xls_sheet]

    return workbook_sheet, workbook

def extract_feature(sheet, row_no):
   # print(row_no)
    #param_values will store the heading of the parameters (e.g Width, Height, Format etc..)
    cell = "A" + str(row_no)
   # print("cell:", cell)
    feature_cell = sheet[cell].value
    feature_cell = "Output/" + str(feature_cell.split(".")[1])
   # print(feature_cell)
    try:
        os.mkdir(feature_cell)
        print("Created ", feature_cell, ": folder")
    except FileExistsError:
        print("folder", feature_cell, "already Exist do you want to continue and replace the data with new one? (y/n)")
        user_input2 = input()
        if user_input2.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return feature_cell

def extract_header(sheet, header_row_number):

    param_values = []
    for cell in sheet[header_row_number]:
        cell_value = cell.value
        param_values.append(cell_value)

   # print(param_values)
    return param_values

def extract_parameters(sheet, next_row, cell_values, output_folder):

    #In XLS we will be using row 4 as the heading of the parameters
    TC_No = sheet.cell(row=next_row, column=1).value
    source_file = 'input_files/input.cfg'
    a = 0
    i = 0
    j = 0

   # print("####")
   # print(cell_values)


    #lines will store the lines made from parsing the table and that we will insert in the table
    lines = []

    #target_text will be holding the cfg section info that in which section line will be apended.
    target_text = []
    avc_flag = 0

    #This loop generates the lines that needs to be updated on cfg file
    for cell in sheet[next_row]:
        value = str(cell_values[i]) + "      =      " + str(cell.value) + " "
        lines.append(value)
        i = i+1
   # print(lines)

    #This block of code generates the cfg files for each testcase
    try:
        os.mkdir(str(output_folder) + "/" + str(TC_No))
        print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
    except FileExistsError:
        print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
        user_input = input()
        if user_input.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()

    #This block compares the target_text with sections inside cfg and gets the lines no. and it will append the line on next line of
    #matching section

    return lines

def parce_error(file_path, error_dict):
    with open(file_path, 'r') as file:
        file_contents = file.read()
        for error_message, error_keywords in error_dict.items():
            for keyword in error_keywords:
                if keyword in file_contents:
                    print(f"Error: {error_message}")
                    error = 1
                    return error



#----------------------------------------------------------------------------------------------------------------------#

parser = argparse.ArgumentParser(description='Testcase_automation_V1.0', add_help=True)

parser.add_argument('-f', '--file', help='Specify a input XLS file for testcases')
parser.add_argument('-s', '--sheet', help='Specify a input XLS file\'s sheet for testcases')
parser.add_argument('-o', '--output', action='store_true', default=False, help='Select this option to store the output file')
parser.add_argument('-tc', '--tc_no', help='Specify individual TC_No that u want to run e.g -tc TC_0001')

args = parser.parse_args()

file_option = args.file
sheet_option = args.sheet
#print(sheet_option)
output_option = args.output

#print(f"file_option: {file_option}")
#print(f"sheet_option: {sheet_option}")
#print(f"output_option: {output_option}")

try:
    os.mkdir("Output")
    print("Created Output folder: Head over to this folder for more TC related information and output files")
except FileExistsError:
    user_input1 = input("Output folder alreaady Exist do you want to continue and replace the data with new one? (y/n)")
    if user_input1.lower() == 'y':
        pass
    else:
        print("Program closing")
        exit()
#xls_file = '/group/siv3/staff/andreis/sibridge/yashl/VCU2/python_script_for_TC/yash_python_scripting/XLS/Encoder_tests.xlsx'
#xls_sheet = 'Temp'
#next_row = 4

CWD = os.getcwd()
#print(CWD)

orignal_xls = args.file
output_xls = "Output/output.xlsx"

#Path of Argon bitstream files
folder_path_argon = "/group/siv3/staff/andreis/vcu_deploy/deliveries/database/input/decoder/complex_dec"
#Path of Fuzz and currupted files path
folder_path_fuzz = "/group/siv3/staff/andreis/vcu_deploy/deliveries/database/input/fuzz_all"
#Path of Conformance HEVC
folder_path_conf = "/group/siv3/staff/andreis/vcu_deploy/deliveries/database/md5sum/xilinx_private/ctrlSW/stream_level/ctrlSW/input/decoder"

#Coping orignal xls into output_xls so we can modify content or put result in output xls
shutil.copy2(orignal_xls, output_xls)

output_sheet, output_workbook = open_workbook(str(output_xls), str(args.sheet))
#It's bit depth flag if no Bitdepth parametre in xls it will remain zero
bd_flag = 0

sheet, new_workbook = open_workbook(str(args.file), str(args.sheet))
#next_row, header_values, Heading_cell = extract_headers(sheet)

#We are iterating over every row for column A
for cell in sheet['A']:
    if sheet_option == "Dec_Conformance":
        no_stream_md5_file = 0

    fill_color = cell.fill.start_color.rgb
   # print(f"The fill color of the cell is: {fill_color}")

   #We are checking for red color cell if red color cell found we are done with all the testcases and program will close
    if sheet[cell.coordinate].fill.start_color.rgb == 'FFFF0000':
        print("\n\n -/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/- \n\n")
        break

    #If there is cell with no black color and it's value is None we skip that row
    if sheet[cell.coordinate].fill.start_color.rgb != 'FF000000' and cell.value is None:
    #    print("Continue", cell.row)
        continue

    #If we find Black color cell we will have feature cell in next row
    if sheet[cell.coordinate].fill.start_color.rgb == 'FF000000':
       # print("#####", cell.row)
        #Enabling flag so in next row we will extract feature name
        extract_feature_flag = 1
     #   print("Black detected at row:", cell.row)
        continue

    if extract_feature_flag == 1:
      #  print("extract_feature condition true")
        feature_folder = extract_feature(sheet, cell.row)
        log_folder = str(CWD) + "/" + str(feature_folder)
       # print(log_folder)
        extract_feature_flag = 0
        #We got the feature name enabling this flag as in next row we will extract the headers of testcase
        extract_header_flag = 1
        continue

    if extract_header_flag == 1:
       # print("Extract headers condition true")
        header_values = extract_header(sheet, cell.row)
       # print("###############", "row", cell.row, len(header_values))
        extract_header_flag = 0
        continue

    if args.tc_no is not None:
#        print("Tc argumenat found")
        if cell.value != args.tc_no:
            continue

#--------------------------------------------------------------------------------------------------------------------    
    parameters = extract_parameters(sheet, cell.row, header_values, log_folder)
#--------------------------------------------------------------------------------------------------------------------    

    substring = "Bitstream"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
       # print("@!@!@!@!@!@!@!@!",final_index)
        bitstream_file = parameters[final_index].split("=")[1]
        bitstream_file = bitstream_file.replace(" ","")
      #  print(bitstream_file)

    substring = "HW MD5sum"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        hw_md5_index = parameters.index(filtered_list[0])

    substring = "Argon Md5Sum"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        md5_index = parameters.index(filtered_list[0])
       # print("@!@!@!@!@!@!@!@!",final_index)
        argon_md5sum = parameters[md5_index].split("=")[1]
        argon_md5sum = argon_md5sum.replace(" ","")
     #   print(argon_md5sum)

    substring = "Category"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        cat_index = parameters.index(filtered_list[0])
       # print("####", cat_index)
       # print("@!@!@!@!@!@!@!@!",final_index)
        cat_folder = parameters[cat_index].split("=")[1]
        cat_folder = cat_folder.replace(" ","")
       # print(cat_folder)
        bd = cat_folder.split("_")[0]
       # print(bd)
        if bd == "main10":
            bd_flag = 1

    substring = "Codec"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        codec_index = parameters.index(filtered_list[0])
       # print("####", cat_index)
       # print("@!@!@!@!@!@!@!@!",final_index)
        codec_value = parameters[codec_index].split("=")[1]
        codec_value = codec_value.replace(" ","")
       # print(codec_value)

    substring = "BitDepth"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        bitdepth_index = parameters.index(filtered_list[0])
       # print("####", cat_index)
       # print("@!@!@!@!@!@!@!@!",final_index)
        bitdepth_value = parameters[bitdepth_index].split("=")[1]
        bitdepth_value = bitdepth_value.replace(" ","")
       # print(bitdepth_value)

    if sheet_option == "Dec_Conformance":
        for root, dirs, files in os.walk(folder_path_conf):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)
        #        print("####COnf File Path: ", file_path)
                stream_md5_file = file_path.rsplit('.', 1)[0] + '.md5'
         #       print(stream_md5_file)
                if os.path.exists(stream_md5_file):
                    pass
                else:
                    print("File not found : ", stream_md5_file)
                    no_stream_md5_file = 1
                try:
                    with open(stream_md5_file, 'r', encoding='utf-8') as file:
                        stream_md5_contents = file.read()
                        stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
          #              print(stream_md5_contents)
                except FileNotFoundError:
                    print("File not found")
                except IOError:
                    print("Error reading file")
        

    if sheet_option == "Dec_Argon":
        final_bitstream_path = folder_path_argon + "/" + cat_folder
        print(final_bitstream_path)
        for root, dirs, files in os.walk(final_bitstream_path):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)
                print("####Argon File Path: ", file_path)

    if sheet_option == "Dec_Fuzz":
       # print("In Elseeeee")
        for root, dirs, files in os.walk(folder_path_fuzz):
            if bitstream_file in dirs:
                fuzz_file_path = os.path.join(root, bitstream_file)
        #        print("fuzz file path: ", fuzz_file_path)
                fuzz_file_list = os.listdir(fuzz_file_path)
                if len(fuzz_file_list) != 0:
                    fuzz_file_name = fuzz_file_list[0]
                    file_path = os.path.join(fuzz_file_path, fuzz_file_name)
                    print(file_path)

    if file_path is None:
        print("Bistream file not found")
       # print(output_file)

    test_case = str(parameters[0].split("=")[1])
    test_case = test_case.replace(" ","") 
    print("Running----------------------", test_case, "-----------------------------------\n\n")

    if "=" in parameters[0]:
        log_file = log_folder + "/" + cell.value + "/" + str(parameters[0].split("=")[1]) + ".txt"
        log_file = log_file.replace(" ","")
   #     print("log file:", log_file)

    with open(log_file, "w") as file:
        current_time = datetime.datetime.now()
        #this the maximum time we will wait for 1 usecase
        deadline = current_time + datetime.timedelta(minutes=5)
        if sheet_option == "Dec_Conformance":
   #         print("###Bitdepth", bitdepth_value)
            if bitdepth_value == "10":
                if codec_value == "AVC":
                    command = "./AL_Decoder.exe -avc -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 10"
                else:
                    command = "./AL_Decoder.exe -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 10"
            else:
                if codec_value == "AVC":
                    command = "./AL_Decoder.exe -avc -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 8"
                else:
                    command = "./AL_Decoder.exe -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 8"

        if sheet_option == "Dec_Argon":
            if bd_flag == 1:
                command = "./AL_Decoder.exe -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 10"
            else:
                command = "./AL_Decoder.exe -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv" + " -bd 8"

        if sheet_option == "Dec_Fuzz":
            command = "./AL_Decoder.exe -in " + str(file_path) + " -out " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv"

        print(command)
        process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        pid = process.pid
        #Polling here until the encoding or decoding is Done
        while process.poll() is None:
            time.sleep(5)
#            if datetime.datetime.now() > deadline:
#                time_failure = 1
#                break

    if sheet_option == "Dec_Conformance":
        md5sum_file = log_folder + "/" + cell.value + "/" + str(parameters[0].split("=")[1]) + "_md5sum.txt"
        md5sum_file = md5sum_file.replace(" ","")
    #   print(md5sum_file)
        with open(md5sum_file, "w") as file:
            command = "md5sum " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv"
            print(command)
            process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
            process.wait()
        try:
            with open(md5sum_file, 'r', encoding='utf-8') as file:
                hw_md5_contents = file.read()
                hw_md5_contents = hw_md5_contents.split(" ")[0]
    #            print("####YUV Md5sum: ", hw_md5_contents)
        except FileNotFoundError:
            print("File not found")
        except IOError:
            print("Error reading file")

    if sheet_option == "Dec_Argon":
        md5sum_file = log_folder + "/" + cell.value + "/" + str(parameters[0].split("=")[1]) + "_md5sum.txt"
        md5sum_file = md5sum_file.replace(" ","")
    #   print(md5sum_file)
        with open(md5sum_file, "w") as file:
            command = "md5sum " + log_folder + "/" + cell.value + "/" + cell.value + ".yuv"
            print(command)
            process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
            process.wait()
        try:
            with open(md5sum_file, 'r', encoding='utf-8') as file:
                hw_md5_contents = file.read()
                hw_md5_contents = hw_md5_contents.split(" ")[0]
  #             print(hw_md5_contents)
        except FileNotFoundError:
            print("File not found")
        except IOError:
            print("Error reading file")

    #if sheet_option == "Dec_Fuzz":
    #    with open(log_file, "r") as file:
    #        fuzz_log_file_content = file.read()
    #        if "No Frames Decoded" in fuzz_log_file_content:
    #            print("###################Match")
    #            no_decoded_frame_flag = 1

    print("\n\n", parameters, "\n\n")

    if sheet_option == "Dec_Conformance":
        try:
            stream_md5_index = header_values.index('Stream MD5sum')
            stream_md5_flag = 0
        except:
            stream_md5_flag = 1

    try:
        index = header_values.index('Result')
        result_flag = 0
    except:
        result_flag = 1

    try:
        hw_md5_index = header_values.index('HW MD5sum')
        hw_md5_result_flag = 0
    except:
        hw_md5_result_flag = 1

    if sheet_option == "Dec_Argon":
        if argon_md5sum == hw_md5_contents:
            md5_flag = 1
        else:
            md5_flag = 0

    if sheet_option == "Dec_Conformance":
        if stream_md5_contents == hw_md5_contents:
            matching_md5_flag = 1
        else:
            matching_md5_flag = 0

    error_flag = parce_error(log_file, error_dict)

    if sheet_option == "Dec_Conformance":
        if error_flag != 1 and result_flag != 1 and matching_md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
            result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
            result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                result_cell.value = hw_md5_contents
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"
            if no_stream_md5_file != 1:
                result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
                result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                 result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                 result_cell.value = hw_md5_contents

    if sheet_option == "Dec_Argon":
        if error_flag != 1 and result_flag != 1 and md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"

    if sheet_option == "Dec_Fuzz":
        if error_flag != 1 and result_flag != 1:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"

    if sheet_option == "Dec_Argon":
        if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
            result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
            result_cell.value = hw_md5_contents
    output_workbook.save(output_xls)

    print("Completed----------------------", test_case, "-----------------------------------\n\n")

