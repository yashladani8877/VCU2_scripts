import random
import inspect
import os
import io
import sys
import openpyxl
import shutil
import fileinput
import subprocess
import time
import datetime
import argparse
import re
import glob
import signal

param_dict = {}
keys_values = {
        'GOP': {
            'GopCtrlMode' : ['ADAPTIVE_GOP', 'DEFAULT_GOP', 'DEFAULT_GOP_B', 'LOW_DELAY_B', 'LOW_DELAY_P', 'PYRAMIDAL_GOP', 'PYRAMIDAL_GOP_B'],
            'Gop.Length'  : ['0', '1', '4', '5', '15', '30', '60', '80', '120', '240', '300', '500', '1000'],
            'Gop.DoubleRef': ['ENABLE', 'DISABLE', 'TRUE', 'FALSE'],
            'Gop.GdrMode' : ['DISABLE', 'GDR_HORIZONTAL', 'GDR_OFF', 'GDR_VERTICAL'],
            'Gop.EnableLT': ['DISABLE', 'ENABLE', 'FALSE', 'TRUE'],
            'Gop.FreqIDR' : ['DISABLE', 'SC_ONLY', 10, 30, 45, 60],
            'Gop.FreqLT'  : [30, 40, 50, 60],
            'Gop.TempDQP' : ['0 0 0 0', '1 2 2 2'],
            'Gop.NumB':{},
            'Gop.FreqRP':['0','1','4','5','7','12','15','23','30','38','45','50']
        },

        'INPUT': {
            'I|Width' : ['1920', '1080', '2048', '2560', '3840', '4096', '7680', '1280', '720', '640', '480', '352', '128'],
            'I|Height': ['1080', '1080', '1440', '2160', '2160'],
            'FrameRate' : ['30', '24', '60', '120', '240'],
            'I|YUVFile':{}, 
            'I|Format' : {},
            'QpTablesFolder' : {},
            'ROIFile' : {}

        },

        'DYNAMIC_INPUT': ['D|Height', 'D|Width', 'D|YUVFile'],

        'OUTPUT': ['BitstreamFile', 'O|CropHeight', 'O|CropPosX', 'O|CropPosY', 'O|CropWidth', 'O|Format', 'RecFile'],

        'RATE_CONTROL': {#'EnableSkip', 'FrameRate', 'IPDelta',
        #'MaxConsecutiveSkip', 'MaxPictureSize.B', 'MaxPictureSize.I', 'MaxPictureSize.P',
        #'MaxPictureSizeInBits', 'MaxPictureSizeInBits.B', 'MaxPictureSizeInBits.I', 'MaxPictureSizeInBits.P', 'MaxPSNR',
        #'MaxQP.B', 'MaxQP.I', 'MaxQP.P', 'MinPSNR', 'MinQP.B', 'MinQP.I',
        #'MinQP.P', 'PBDelta', 'ScnChgResilience', 'SCPrevention', 'UseGoldenRef',
            'RateCtrlMode' : ['CAPPED_VBR','LOW_LATENCY','PLUGIN','VBR','CBR','CONST_QP'],
            'MaxBitRate' : ['4000','960000','720000','600000','480000','400080','360000','300000','240000','19920','160080'],
            'BitRate' : ['80','106','85','128','256','192','214','320','384','80000','100000','120000','150000','160080','180000','199920','200000','230000','240000','270000','300000','350000','360000','400080','450000','460000','480000','520000','600000','700000','720000','960000'],
            'MaxQP' : ['AUTO','0','24','36','51'],
            'MinQP' : ['AUTO','0','2','8','15','24','30','36','46','51'],
            'SliceQP' : ['AUTO','0','2','8','15','24','30','36','46','51'],
            'CPBSize' : {},
            'InitialDelay' : {},
            'MaxPictureSize' : {}
        },

        'SETTINGS': {
            'ChromaMode' : ['CHROMA_4_0_0', 'CHROMA_4_2_0', 'CHROMA_4_2_2', 'CHROMA_4_4_4'],
            'BitDepth'   : ['8', '10', '12'],
            'QPCtrlMode' : ['ADAPTIVE_AUTO_QP','AUTO_QP','LOAD_QP | RELATIVE_QP','ROI_QP','UNIFORM_QP','LOAD_QP'],
            'VideoMode' : ['PROGRESSIVE','INTERLACED_BOTTOM','INTERLACED_TOP'],
            'EnableSEI' : ['SEI_ALL','SEI_ATC','SEI_BP','SEI_CLL','SEI_MDCV','SEI_NONE','SEI_PT','SEI_RP','SEI_ST2094_10','SEI_ST2094_40'],
            'CabacInit' : ['DISABLE','DISABLE','0','1'],
            'ScalingList' : ['DEFAULT','FLAT'],
            'UseL2C' : ['DISABLE','ENABLE','FALSE','TRUE'],
            'PicCbQpOffset' : ['-12','-5','-3','-1','0','2','7','12'],
            'PicCrQpOffset' : ['-12','-5','-3','-1','0','2','7','12'],
            'SubframeLatency' : ['DISABLE','ENABLE','FALSE','TRUE'],
            'Profile':{},
            'SrcFormat':{}
        },

        'RUN': {
            'Loop' : ['ENABLE', 'DISABLE', 'TRUE', 'FALSE'],
            'MaxPicture':{}
        }
}

for key, values in keys_values.items():
    param_dict[key] = values

width_height_mapping = {
        '1920' : '1080',
        '1080' : '1080',
        '2048' : '1080',
        '2560' : '1440',
        '3840' : '2160',
        '4096' : '2160',
        '7680' : '4320',
        '1280' :  '720',
        '720'  :  '576',
        '640'  :  '480',
        '480'  :  '360',
        '352'  :  '240',
        '128'  :  '64'
}

format_mapping = {
        'CHROMA_4_0_0' : {
            '8' : ['T5m8','T6m8','Y800'],
            '10': ['T5mA','T6mA','Y010'],
            '12': ['T5mC','T6mC','Y012']
        },
        'CHROMA_4_2_0' : {
            '8' : ['T508','T608','I420','NV12'],
            '10': ['T50A','T60A','I0AL','P010'],
            '12': ['T50C','T60C','I0CL','P012']
        },
        'CHROMA_4_2_2' : {
            '8' : ['T528','T628','I422','NV16'],
            '10': ['T52A','T62A','I2AL','P210'],
            '12': ['T52C','T62C','I2CL','P212']
        },
        'CHROMA_4_4_4' : {
            '8' : ['T648','T548','I444'],
            '10': ['T54A','T64A','I4AL'],
            '12': ['T54C','T64C','I4CL']
        }
}
profile_mapping = {
        'CHROMA_4_0_0' : {
            '8' : ['HEVC_MONO','AVC_HIGH'],
            '10': ['HEVC_MONO10','AVC_HIGH10'],
            '12': ['HEVC_MAIN12','AVC_HIGH10']
        },
        'CHROMA_4_2_0' : {
            '8' : ['HEVC_MAIN','AVC_HIGH'],
            '10': ['HEVC_MAIN10','AVC_HIGH10'],
            '12': ['HEVC_MAIN12','AVC_HIGH10']
        },
        'CHROMA_4_2_2' : {
            '8' : ['HEVC_MAIN_422','AVC_HIGH_422'],
            '10': ['HEVC_MAIN_422_10','AVC_HIGH_422'],
            '12': ['HEVC_MAIN_422_12','AVC_HIGH_422']
        },
        'CHROMA_4_4_4' : {
            '8' : ['HEVC_MAIN_444','AVC_HIGH_444_PRED'],
            '10': ['HEVC_MAIN_444_10','AVC_HIGH_444_PRED'],
            '12': ['HEVC_MAIN_444_12','AVC_HIGH_444_PRED']
        }
}

gop_numb_mapping = {
        'DEFAULT_GOP': {
            '0' : ['0'],
            '1' : ['0'],
            '4' : ['1', '0'],
            '5' : ['0', '4'],
            '15': ['0', '2', '4'],
            '30': ['0', '1', '2', '4', '5'],
            '60': ['0', '1', '2', '3', '4'],
            '80': ['0', '1', '3', '7'],
            '120': ['0', '1', '4', '5'],
            '240': ['0', '1', '2', '3', '4'],
            '300': ['0', '1', '2', '5', '9'],
            '500': ['0', '1', '4', '9'],
            '1000': ['0', '1', '4', '9',]
            },
        'PYRAMIDAL_GOP': {
            '0' : ['0'],
            '1' : ['0'],
            '4' : ['0', '3'],
            '5' : ['0'],
            '15': ['0', '3'],
            '30': ['5'],
            '60': ['3', '5'],
            '80': ['3', '7','15'],
            '120': ['3', '5', '7'],
            '240': ['3', '5', '7','15'],
            '300': ['3', '5', '7'],
            '500': ['3', '5', '7'],
            '1000': ['3', '5', '7']
            },
        'PYRAMIDAL_GOP_B': {
            '0' : ['0'],
            '1' : ['0'],
            '4' : ['0', '3'],
            '5' : ['0'],
            '15': ['0', '3'],
            '30': ['5'],
            '60': ['3', '5'],
            '80': ['3', '7','15'],
            '120': ['3', '5', '7'],
            '240': ['3', '5', '7','15'],
            '300': ['3', '5', '7'],
            '500': ['3', '5', '7'],
            '1000': ['3', '5', '7']
            }    
        }

maxpicture_maping = {
        'ENABLE' : ['60', '70', '80', '90', '100', '110', '120'],
        'TRUE'   : ['60', '70', '80', '90', '100', '110', '120'],
        'DISABLE': ['ALL'],
        'FALSE'  : ['ALL']
        }
qpfile_maping = {
        '3840': {
            'avc':['input_files/QPs_4k_avc', 'input_files/QPs_4k_avc_intra'],
            'hevc':['input_files/QPs_4k_hevc', 'input_files/QPs_4k_hevc_intra']
            },
        '1920':{
            'avc':['input_files/QPs_1080_avc', 'input_files/QPs_1080_avc_intra'],
            'hevc':['input_files/QPs_1080_hevc', 'input_files/QPs_1080_hevc_intra']
            }
        }


error_dict = {}

error_msg = {
        'Assertion \'0\' failed': ['Assertion|failed', 'Assertion', 'assertion'],
        'Encoder Resource Error': ['Failed to create Encoder'],
        'CMA allocation Error': ['Cannot allocate memory'],
        'No QP File found': ['No QP File found'],
        'Unknown identifire please check property name': ['unknown identifier'],
        'I/p YUVFile not found': ['Exception caught: Can\'t open file for reading'],
        'Get higher Profile to support the usecase': ['getHevcMinimumProfile: Assertion \`0\' failed'],
        'Segmentation fault Error':['Segmentation fault'],
        'Invalid Parameter in settings':['errors(s). Invalid settings'],
        'Exception caught Error': ['Exception caught'],
        'Error in ctrlsw app': ['Error', 'error', 'ERROR']
}

for key, values in error_msg.items():
    error_dict[key] = values


#param_dict[key] = values

def open_workbook(xls_file, xls_sheet):

    #open workbook
    workbook = openpyxl.load_workbook(xls_file)

    #Select the sheet by Index or name
    workbook_sheet = workbook[xls_sheet]

    return workbook_sheet, workbook

def extract_feature(sheet, row_no):
   # print(row_no)
    #param_values will store the heading of the parameters (e.g Width, Height, Format etc..)
    cell = "A" + str(row_no)
   # print("cell:", cell)
    feature_cell = sheet[cell].value
    feature_cell = "Output/" + "Random_folder"
   # print(feature_cell)
    try:
        os.mkdir(feature_cell)
        print("Created ", feature_cell, ": folder")
    except FileExistsError:
        print("folder", feature_cell, "already Exist do you want to continue and replace the data with new one? (y/n)")
        user_input2 = input()
        if user_input2.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return feature_cell

def extract_header(sheet, header_row_number):

    param_values = []
    for cell in sheet[header_row_number]:
        cell_value = cell.value
        param_values.append(cell_value)

    return param_values

def generate_parameters(sheet, next_row, cell_values, output_folder):

   # print(f"next row : {next_row}")
    #In XLS we will be using row 4 as the heading of the parameters
    #TC_No = sheet.cell(row=next_row, column=1).value
    source_file = 'input_files/input.cfg'
    a = 0
    i = 0
    j = 0
    skip = 0
    x  = next_row - 2
    TC_No = "TC_" + str(x)

    #lines will store the lines made from parsing the table and that we will insert in the table
    lines = []
    random_param_array = []

    #target_text will be holding the cfg section info that in which section line will be apended.
    target_text = []
    avc_flag = 0

    for cell in sheet[next_row]:
        for key, values in param_dict.items():
            if cell_values[i] in values:
                #This if condition checks for same named parameters and put them in according sections
                if "|" in cell_values[i]:
                    a = 1
                    split_parts = cell_values[i].split("|")
                    cell_values[i] = cell_values[i].split("|")[1]
                    if cell_values[i] == "Width":
                        width = cell.value
                    if cell_values[i] == "Height":
                        height = cell.value
                    if cell_values[i] == "Format":
                        Format = cell.value
                    if cell_values[i] == "YUVFile":
                        YUVFile = cell.value
                target_text.append(key)
                break
        value = str(cell_values[i]) + "      =      " + str(cell.value) + " "
        lines.append(value)
        if a == 1:
            cell_values[i] = "|".join(split_parts)
            a = 0
        i = i+1


    #This block of code generates the cfg files for each testcase
    if skip != 1:
        try:
            os.mkdir(str(output_folder) + "/" + str(TC_No))
            print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
        except FileExistsError:
            print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
            user_input = input()
            if user_input.lower() == 'y':
                pass
            else:
                print("Program closing")
                exit()
    if skip != 1:
        destination_file = f'{output_folder}/{TC_No}/input_{TC_No}.cfg'
        shutil.copy2(source_file, destination_file)

    for element in lines:
        random_param = element.split('=')[0]
        random_param = random_param.replace(" ","")
        if random_param == "ChromaMode":
            if args.chroma_mode is None:
                random_chromamode_value = random.choice(keys_values['SETTINGS']['ChromaMode'])
            else:
                random_chromamode_value = args.chroma_mode
            for i in range(len(lines)):
                if lines[i].startswith('ChromaMode'):
                    lines[i] = str(random_param) + "   =   " + str(random_chromamode_value)
                    break
            continue
        elif random_param == "Width":
            if args.width is None:
                random_width_value = random.choice(keys_values['INPUT']['I|Width'])
            else:
                random_width_value = args.width
            for i in range(len(lines)):
                if lines[i].startswith('Width'):
                    lines[i] = str(random_param) + "   =   " + str(random_width_value)
                    break
            continue
        elif random_param == "Height":
            if args.height is None:
                random_height_value = width_height_mapping.get(random_width_value)
            else:
                random_height_value = args.height
            for i in range(len(lines)):
                if lines[i].startswith('Height'):
                    lines[i] = str(random_param) + "   =   " + str(random_height_value)
                    break
            continue
        elif random_param == "FrameRate":
            if args.framerate is None:
                random_framerate_value = random.choice(keys_values['INPUT']['FrameRate'])
            else:
                random_framerate_value = args.framerate
            for i in range(len(lines)):
                if lines[i].startswith('FrameRate'):
                    lines[i] = str(random_param) + "   =   " + str(random_framerate_value)
                    break
            continue
        elif random_param == "BitDepth":
            if args.bitdepth is None:
                random_bitdepth_value = random.choice(keys_values['SETTINGS']['BitDepth'])
            else:
                random_bitdepth_value = args.bitdepth
            for i in range(len(lines)):
                if lines[i].startswith('BitDepth'):
                    lines[i] = str(random_param) + "   =   " + str(random_bitdepth_value)
                    break
            continue
        elif random_param == "Format":
            if args.format is None:
                random_format_value = random.choice(format_mapping[random_chromamode_value][random_bitdepth_value])
            else:
                random_format_value = args.format
            for i in range(len(lines)):
                if lines[i].startswith('Format'):
                    lines[i] = str(random_param) + "   =   " + str(random_format_value)
                    break
            continue
        elif random_param == "SrcFormat":
            if random_format_value in ('T508','T50A','T50C','T528','T52A','T52C','T54A','T54C','T5m8','T5mA','T5mC','T648'):
                random_srcformat_value = "TILE_32x4"
            elif random_format_value in ('T608','T60A','T60C','T628','T62A','T62C','T64A','T64C','T6m8','T6mA','T6mC','T548'):
                random_srcformat_value = "TILE_64x4"
            else:
                random_srcformat_value = "NVX"
            for i in range(len(lines)):
                if lines[i].startswith('SrcFormat'):
                    lines[i] = str(random_param) + "   =   " + str(random_srcformat_value)
                    break
            continue
        elif random_param == "Profile":
            while 1:
                random_profile_value = random.choice(profile_mapping[random_chromamode_value][random_bitdepth_value])
                if random_profile_value in ('HEVC_MAIN','HEVC_MAIN10','HEVC_MAIN10_INTRA','HEVC_MAIN12','HEVC_MAIN_422',
                        'HEVC_MAIN_422_10','HEVC_MAIN_422_10_INTRA','HEVC_MAIN_422_12','HEVC_MAIN_444','HEVC_MAIN_444_10',
                        'HEVC_MAIN_444_10_INTRA','HEVC_MAIN_444_12','HEVC_MAIN_444_INTRA','HEVC_MAIN_444_STILL','HEVC_MAIN_INTRA',
                        'HEVC_MAIN_STILL','HEVC_MONO','HEVC_MONO10') :
                    codec = 'hevc'
                else :
                    codec = 'avc'
                if codec == args.codec or args.codec is None:
                    break
            for i in range(len(lines)):
                if lines[i].startswith('Profile'):
                    lines[i] = str(random_param) + "   =   " + str(random_profile_value)
                    break
            continue
        elif random_param == "VideoMode":
            if codec == 'hevc' :
                random_videomode_value = random.choice(keys_values['SETTINGS']['VideoMode'])
                for i in range(len(lines)):
                    if lines[i].startswith('VideoMode'):
                        lines[i] = str(random_param) + "   =   " + str(random_videomode_value)
                        break
                continue
            else :
                for i in range(len(lines)):
                    if lines[i].startswith('VideoMode'):
                        lines[i] = str(random_param) + "   =   " + "PROGRESSIVE"
                        break
                continue
        elif random_param == "EnableSEI":
            random_enablesei_value = random.choice(keys_values['SETTINGS']['EnableSEI'])
            for i in range(len(lines)):
                if lines[i].startswith('EnableSEI'):
                    lines[i] = str(random_param) + "   =   " + str(random_enablesei_value)
                    break
            continue
        elif random_param == "CabacInit":
            random_cabacinit_value = random.choice(keys_values['SETTINGS']['CabacInit'])
            for i in range(len(lines)):
                if lines[i].startswith('CabacInit'):
                    lines[i] = str(random_param) + "   =   " + str(random_cabacinit_value)
                    break
            continue
        elif random_param == "ScalingList":
            random_scalinglist_value = random.choice(keys_values['SETTINGS']['ScalingList'])
            for i in range(len(lines)):
                if lines[i].startswith('ScalingList'):
                    lines[i] = str(random_param) + "   =   " + str(random_scalinglist_value)
                    break
            continue
        elif random_param == "PicCbQpOffset":
            random_piccbqpoffset_value = random.choice(keys_values['SETTINGS']['PicCbQpOffset'])
            for i in range(len(lines)):
                if lines[i].startswith('PicCbQpOffset'):
                    lines[i] = str(random_param) + "   =   " + str(random_piccbqpoffset_value)
                    break
            continue
        elif random_param == "PicCrQpOffset":
            random_piccrqpoffset_value = random.choice(keys_values['SETTINGS']['PicCrQpOffset'])
            for i in range(len(lines)):
                if lines[i].startswith('PicCrQpOffset'):
                    lines[i] = str(random_param) + "   =   " + str(random_piccrqpoffset_value)
                    break
            continue
        elif random_param == "SubframeLatency":
            random_subframelatency_value = random.choice(keys_values['SETTINGS']['SubframeLatency'])
            for i in range(len(lines)):
                if lines[i].startswith('SubframeLatency'):
                    lines[i] = str(random_param) + "   =   " + str(random_subframelatency_value)
                    break
            continue
        elif random_param == "UseL2C":
            random_usel2c_value = random.choice(keys_values['SETTINGS']['UseL2C'])
            for i in range(len(lines)):
                if lines[i].startswith('UseL2C'):
                    lines[i] = str(random_param) + "   =   " + str(random_usel2c_value)
                    break
            continue
        elif random_param == "GopCtrlMode":
            random_gopctrlmode_value = random.choice(keys_values['GOP']['GopCtrlMode'])
            for i in range(len(lines)):
                if lines[i].startswith('GopCtrlMode'):
                    lines[i] = str(random_param) + "   =   " + str(random_gopctrlmode_value)
                    break
            continue
        elif random_param == "Gop.Length":
            random_goplength_value = random.choice(keys_values['GOP']['Gop.Length'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.Length'):
                    lines[i] = str(random_param) + "   =   " + str(random_goplength_value)
                    break
            continue
        elif random_param == "Gop.NumB":
            if random_gopctrlmode_value != "ADAPTIVE_GOP" and random_gopctrlmode_value != "LOW_DELAY_P" and random_gopctrlmode_value != "LOW_DELAY_B" and random_gopctrlmode_value != "DEFAULT_GOP_B":
                random_gopnumb_value = random.choice(gop_numb_mapping[random_gopctrlmode_value][random_goplength_value])
                for i in range(len(lines)):
                    if lines[i].startswith('Gop.NumB'):
                        lines[i] = str(random_param) + "   =   " + str(random_gopnumb_value)
                        break
                continue
            else:
                for i in range(len(lines)):
                    if lines[i].startswith('Gop.NumB'):
                        lines[i] = str(random_param) + "   =   0 " 
                        break
                continue
        elif random_param == "Gop.FreqRP":
            random_freqrp_value = random.choice(keys_values['GOP']['Gop.FreqRP'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.FreqRP'):
                    lines[i] = str(random_param) + "   =   " + str(random_freqrp_value)
                    break
            continue
        elif random_param == "Loop":
            random_loop_value = random.choice(keys_values['RUN']['Loop'])
            for i in range(len(lines)):
                if lines[i].startswith('Loop'):
                    lines[i] = str(random_param) + "   =   " + str(random_loop_value)
                    break
            continue
        elif random_param == "MaxPicture":
            random_maxpicture_value = random.choice(maxpicture_maping[random_loop_value])
            for i in range(len(lines)):
                if lines[i].startswith('MaxPicture'):
                    lines[i] = str(random_param) + "   =   " + str(random_maxpicture_value)
                    break
            continue
        elif random_param == "Gop.DoubleRef":
            random_gopdoubleref_value = random.choice(keys_values['GOP']['Gop.DoubleRef'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.DoubleRef'):
                    lines[i] = str(random_param) + "   =   " + str(random_gopdoubleref_value)
                    break
            continue
        elif random_param == "Gop.GdrMode":
            if random_gopctrlmode_value == "LOW_DELAY_B" or random_gopctrlmode_value == "LOW_DELAY_P":
                random_gopgdrmode_value = random.choice(keys_values['GOP']['Gop.GdrMode'])
                for i in range(len(lines)):
                    if lines[i].startswith('Gop.GdrMode'):
                        lines[i] = str(random_param) + "   =   " + str(random_gopgdrmode_value)
                        break
                continue
            else:
                for i in range(len(lines)):
                    if lines[i].startswith('Gop.GdrMode'):
                        lines[i] = str(random_param) + "   =   " + "DISABLE"
                        break
                continue

        elif random_param == "Gop.EnableLT":
            random_gopenablelt_value = random.choice(keys_values['GOP']['Gop.EnableLT'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.EnableLT'):
                    lines[i] = str(random_param) + "   =   " + str(random_gopenablelt_value)
                    break
            continue
        elif random_param == "Gop.FreqIDR":
            random_gopfreqidr_value = random.choice(keys_values['GOP']['Gop.FreqIDR'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.FreqIDR'):
                    lines[i] = str(random_param) + "   =   " + str(random_gopfreqidr_value)
                    break
            continue
        elif random_param == "Gop.FreqLT":
            random_gopfreqlt_value = random.choice(keys_values['GOP']['Gop.FreqLT'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.FreqLT'):
                    lines[i] = str(random_param) + "   =   " + str(random_gopfreqlt_value)
                    break
            continue
        elif random_param == "Gop.TempDQP":
            random_goptempdqp_value = random.choice(keys_values['GOP']['Gop.TempDQP'])
            for i in range(len(lines)):
                if lines[i].startswith('Gop.TempDQP'):
                    lines[i] = str(random_param) + "   =   " + str(random_goptempdqp_value)
                    break
            continue
        elif random_param == "QPCtrlMode":
            random_qpctrlmode_value = random.choice(keys_values['SETTINGS']['QPCtrlMode'])
            while random_qpctrlmode_value == "LOAD_QP" or random_qpctrlmode_value == 'LOAD_QP | RELATIVE_QP':
                if random_width_value == '3840' or random_width_value == '1920' :
                    break
                random_qpctrlmode_value = random.choice(keys_values['SETTINGS']['QPCtrlMode'])
            for i in range(len(lines)):
                if lines[i].startswith('QPCtrlMode'):
                    lines[i] = str(random_param) + "   =   " + str(random_qpctrlmode_value)
                    break
            continue
        elif random_param == "RateCtrlMode":
            random_ratectrlmode_value = random.choice(keys_values['RATE_CONTROL']['RateCtrlMode'])
            for i in range(len(lines)):
                if lines[i].startswith('RateCtrlMode'):
                    lines[i] = str(random_param) + "   =   " + str(random_ratectrlmode_value)
                    break
            continue
        elif random_param == "MaxBitRate":
            if random_ratectrlmode_value != 'CBR':          
                random_maxbitrate_value = random.choice(keys_values['RATE_CONTROL']['MaxBitRate'])
                for i in range(len(lines)):
                    if lines[i].startswith('MaxBitRate'):
                        lines[i] = str(random_param) + "   =   " + str(random_maxbitrate_value)
                        break
                continue
            else:
                for i in range(len(lines)):
                    if lines[i].startswith('MaxBitRate'):
                        lines[i] = str(random_param) + "   =   4000 "
                        break
                continue
        elif random_param == "BitRate":
            random_bitrate_value = random.choice(keys_values['RATE_CONTROL']['BitRate'])

            if random_ratectrlmode_value != 'CBR':
                while random_bitrate_value >= random_maxbitrate_value:
                    random_bitrate_value = random.choice(keys_values['RATE_CONTROL']['BitRate'])
            for i in range(len(lines)):
                if lines[i].startswith('BitRate'):
                    lines[i] = str(random_param) + "   =   " + str(random_bitrate_value)
                    break
            continue
        elif random_param == "MaxQP":
            random_maxqp_value = random.choice(keys_values['RATE_CONTROL']['MaxQP'])
            for i in range(len(lines)):
                if lines[i].startswith('MaxQP'):
                    lines[i] = str(random_param) + "   =   " + str(random_maxqp_value)
                    break
            continue
        elif random_param == "MinQP":
            random_minqp_value = random.choice(keys_values['RATE_CONTROL']['MinQP'])
            while random_minqp_value > random_maxqp_value :
                random_minqp_value = random.choice(keys_values['RATE_CONTROL']['MinQP'])
            for i in range(len(lines)):
                if lines[i].startswith('MinQP'):
                    lines[i] = str(random_param) + "   =   " + str(random_minqp_value)
                    break
            continue
        elif random_param == "SliceQP":
            random_sliceqp_value = random.choice(keys_values['RATE_CONTROL']['SliceQP'])
            for i in range(len(lines)):
                if lines[i].startswith('SliceQP'):
                    lines[i] = str(random_param) + "   =   " + str(random_sliceqp_value)
                    break
            continue
        elif random_param == "CPBSize":
            random_cpb_value = random.uniform(0.00, 35555.55)
            random_cpbsize_value = round(random_cpb_value, 2)
            for i in range(len(lines)):
                if lines[i].startswith('CPBSize'):
                    lines[i] = str(random_param) + "   =   " + str(random_cpbsize_value)
                    break
            continue
        elif random_param == "InitialDelay":
            random_initial_value = random.uniform(0.00, 35555.55)
            while random_initial_value >= random_cpbsize_value :
                random_initial_value = random.uniform(0.00, 35555.55)
            random_initialdelay_value = round(random_initial_value, 2)
            for i in range(len(lines)):
                if lines[i].startswith('InitialDelay'):
                    lines[i] = str(random_param) + "   =   " + str(random_initialdelay_value)
                    break
            continue
        elif random_param == "MaxPictureSize":
            random_maxpicture_value = random.randint(0, 4294967295)
            for i in range(len(lines)):
                if lines[i].startswith('MaxPictureSize'):
                    lines[i] = str(random_param) + "   =   " + str(random_maxpicture_value)
                    break
            continue
        elif random_param == "QpTablesFolder":
            if random_qpctrlmode_value == 'LOAD_QP':
                random_qpfile_value = random.choice(qpfile_maping[random_width_value][codec]) 
                for i in range(len(lines)):
                    if lines[i].startswith('QpTablesFolder'):
                        lines[i] = str(random_param) + "   =   " + str(random_qpfile_value) 
                        break
                continue
            elif random_qpctrlmode_value == 'LOAD_QP | RELATIVE_QP':
                random_qpfile_value = random.choice(qpfile_maping[random_width_value][codec]) 
                for i in range(len(lines)):
                    if lines[i].startswith('QpTablesFolder'):
                        lines[i] = str(random_param) + "   =   " + str(random_qpfile_value) + "_rel" 
                        break
                continue
            else:
                for i in range(len(lines)):
                    if lines[i].startswith('QpTablesFolder'):
                        lines[i] = "#" + str(random_param) + "   =   " + " " 
                        break
                continue
        elif random_param == "ROIFile":
            if random_qpctrlmode_value == 'ROI_QP':
                for i in range(len(lines)):
                    if lines[i].startswith('ROIFile'):
                        lines[i] = str(random_param) + "   =   " + "input_files/ROI.txt"
                        break
                continue
            else:
                for i in range(len(lines)):
                    if lines[i].startswith('ROIFile'):
                        lines[i] = "#" + str(random_param) + "   =   " + " "
                        break
                continue
        elif random_param == "YUVFile":
            YUV_Folder = "/everest/ssw_multimedia_bkup/VCU2/video_YUV/Crowd_Run_" + str(random_width_value) + "_" + str(random_height_value)
            search_pattern = f'*_{random_format_value}.*'
            random_yuvfile_value = glob.glob(f'{YUV_Folder}/{search_pattern}')
            for i in range(len(lines)):
                if lines[i].startswith('YUVFile'):
                    lines[i] = str(random_param) + "   =   " + str(random_yuvfile_value[0])
                    break
            continue
        elif random_param == "BitstreamFile":
            for i in range(len(lines)):
                if lines[i].startswith('BitstreamFile'):
                    if codec != 'hevc':
                        lines[i] = str(random_param) + "   =   " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".avc"
                    else:
                        lines[i] = str(random_param) + "   =   " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".hevc"
                    break
            continue
        elif random_param == "TC_No":
            for i in range(len(lines)):
                if lines[i].startswith('TC_No'):
                    lines[i] = str(random_param) + "   =   " + str(TC_No)
                    break
            continue
        elif random_param == "Error":
            for i in range(len(lines)):
                if lines[i].startswith('Error'):
                    lines[i] = str(random_param) + "   =   " + " "
                    break
            continue


    #This block compares the target_text with sections inside cfg and gets the lines no. and it will append the line on next line of
    #matching section
    i = 1
    for k in range(len(lines)-3):
        if skip != 1:
            with open(destination_file, 'r') as file:
                for line_num, line in enumerate(file, 1):
                    if target_text[j] in line:
                        final_line = line_num

                with open(destination_file, 'r') as file:
                        line1 = file.readlines()

                if final_line >= 1 and final_line <= len(line1) + 1:
                        line1.insert(final_line, lines[i] + '\n')

                with open(destination_file, 'w') as file:
                            file.writelines(line1)
                j = j+1
                i = i+1

    return lines

def parce_error(file_path, error_dict):
    with open(file_path, 'r') as file:
        file_contents = file.read()
        for error_message, error_keywords in error_dict.items():
            for keyword in error_keywords:
                if keyword in file_contents:
                    error = 1
                    print(f"Error: {error_message}")
                    return error, error_message
    return 0, None



#----------------------------------------------------------------------------------------------------------------------#

parser = argparse.ArgumentParser(description='Testcase_automation_V1.0', add_help=True)

parser.add_argument('-f', '--file', help='Specify a input XLS file for testcases')
parser.add_argument('-s', '--sheet', help='Specify a input XLS file\'s sheet for testcases')
parser.add_argument('-o', '--output', action='store_true', default=False, help='Select this option to store the output file')
parser.add_argument('--width', help='Specify individual witdh that u want to run e.g --weight 1920')
parser.add_argument('--height', help='Specify individual height that u want to run e.g --height 1080')
parser.add_argument('--chroma_mode', help='Specify chroma-mode (CHROMA_MONO, CHROMA_4_0_0, CHROMA_4_2_0, CHROMA_4_2_2, CHROMA_4_4_4)')
parser.add_argument('--bitdepth', help='Specify bitdepth (8, 10, 12')
parser.add_argument('--framerate', help='Specify individual Frame Rate that u want to run [ 0 to 240 ]')
parser.add_argument('--format', help='Specifies YUV input format (I420, IYUV, YV12, NV12, Y800, Y010, P010, I0AL ...)')
parser.add_argument('--codec', help='Specify codec that you want to run (avc or hevc)')

args = parser.parse_args()

file_option = args.file
sheet_option = args.sheet
output_option = args.output


try:
    os.mkdir("Output")
    print("Created Output folder: Head over to this folder for more TC related information and output files")
except FileExistsError:
    user_input1 = input("Output folder alreaady Exist do you want to continue and replace the data with new one? (y/n)")
    if user_input1.lower() == 'y':
        pass
    else:
        print("Program closing")
        exit()

CWD = os.getcwd()

orignal_xls = args.file
output_xls = "Output/output.xlsx"

shutil.copy2(orignal_xls, output_xls)

output_sheet, output_workbook = open_workbook(str(output_xls), str(args.sheet))

sheet, new_workbook = open_workbook(str(args.file), str(args.sheet))

index = 0
current_row = 4
current_column =1 
while 1 :
    index = index + 1
    cell = sheet.cell(row=index, column=1)
    time_failure = 0
    if sheet[cell.coordinate].fill.start_color.rgb == 'FF000000':
        #Enabling flag so in next row we will extract feature name
        extract_feature_flag = 1
     #   print("Black detected at row:", cell.row)
        continue
    if extract_feature_flag == 1:
        feature_folder = extract_feature(sheet, cell.row)
        log_folder = str(CWD) + "/" + str(feature_folder)
        extract_feature_flag = 0
        #We got the feature name enabling this flag as in next row we will extract the headers of testcase
        extract_header_flag = 1
        continue
    if extract_header_flag == 1:
        header_values = extract_header(sheet, cell.row)
        result_index = header_values.index('Result')
        bitstream_index = header_values.index('BitstreamFile')
        extract_header_flag = 0
    
    parameters = generate_parameters(sheet, cell.row, header_values, log_folder)
    print(parameters)
    substring = "Result"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
        output_string = parameters[final_index].split("=")[1]
        output_string = output_string.replace(" ","")
        if output_string == "PASS":
            continue

    test_case = str(parameters[0].split("=")[1])
    test_case = test_case.replace(" ","") 
    print("Running----------------------", test_case, "-----------------------------------\n\n")
    if "=" in parameters[0]:
        log_file = log_folder + "/" + str(parameters[0].split("=")[1]) + "/" + str(parameters[0].split("=")[1]) + ".txt"
        log_file = log_file.replace(" ","")
#        print("log file:", log_file)
    with open(log_file, "w") as file:
        current_time = datetime.datetime.now()
        #this the maximum time we will wait for 1 usecase
        deadline = current_time + datetime.timedelta(minutes=180)
        meminfo_command = "cat /proc/meminfo"
        process = subprocess.Popen(meminfo_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        process.wait()
        command = "./AL_Encoder.exe -cfg " + str(log_folder) + "/" + str(test_case) + "/" + "input_" + str(test_case) + ".cfg"
        print(command)
        #process = subprocess.Popen(command, shell=True, stdout=file, text=True)
        process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
#        process = subprocess.Popen("ls -lrt", shell=True, stdout=file, stderr=subprocess.PIPE, text=True)
        pid = process.pid
        #Polling here until the encoding or decoding is Done
        while process.poll() is None:
            time.sleep(5)
            if datetime.datetime.now() > deadline:
                time_failure = 1
                try:
                    os.kill(pid, signal.SIGTERM)
                    print("Due to hang, ", test_case, "is failed and process : ", pid, "is killed")
                    break
                except OSError:
                    print("Could not kill the process : ", pid)
#                   break
        process = subprocess.Popen(meminfo_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        process.wait()
    error_flag, error_msg = parce_error(log_file, error_dict)

    for header in header_values :
        if current_column == (bitstream_index+1) :
            output_sheet.cell(row=current_row,column=current_column,value=str(parameters[current_column-1].split("/")[-1]))
        elif current_column == (result_index +1) :
            if error_flag != 1 and time_failure != 1 :
                output_sheet.cell(row=current_row,column=current_column,value='PASS')
                print("Testcase_No : ",test_case, "Encoded Succecfully")
            else :
                output_sheet.cell(row=current_row,column=current_column,value='FAIL')
        elif current_column == (result_index + 2):
            output_sheet.cell(row=current_row,column=current_column,value=str(error_msg))
        else :
            output_sheet.cell(row=current_row,column=current_column,value=str(parameters[current_column-1].split("=")[1]))
        current_column += 1
    current_row += 1
    current_column = 1
    output_workbook.save(output_xls)
    print("Completed----------------------", test_case, "-----------------------------------\n\n")

