import os
import io
import sys
import openpyxl
import shutil
import fileinput
import subprocess
import time
import datetime
import argparse
import re
import glob

param_dict = {}
keys_values = {
        'GOP': ['Gop.DoubleRef', 'Gop.EnableLT', 'Gop.FreqIDR', 'Gop.FreqLT', 'Gop.FreqRP', 'Gop.GdrMode', 'Gop.Length',
        'Gop.NumB', 'Gop.TempDQP', 'Gop.WriteAVCHeaderSVC', 'GopCtrlMode'],

        'INPUT': ['CmdFile', 'I|CropHeight', 'I|CropPosX', 'I|CropPosY', 'I|CropWidth', 'I|Format', 'I|FrameRate', 'GMVFile', 'HDRFile',
        'I|Height', 'MapFile', 'QpTablesFolder', 'ROIFile', 'I|Width', 'I|YUVFile'],

        'DYNAMIC_INPUT': ['D|Height', 'D|Width', 'D|YUVFile'],

        'OUTPUT': ['BitstreamFile', 'O|CropHeight', 'O|CropPosX', 'O|CropPosY', 'O|CropWidth', 'O|Format', 'RecFile'],

        'RATE_CONTROL': ['BitRate', 'CPBSize', 'EnableSkip', 'FrameRate', 'InitialDelay', 'IPDelta', 'MaxBitRate',
        'MaxConsecutiveSkip', 'MaxPictureSize', 'MaxPictureSize.B', 'MaxPictureSize.I', 'MaxPictureSize.P',
        'MaxPictureSizeInBits', 'MaxPictureSizeInBits.B', 'MaxPictureSizeInBits.I', 'MaxPictureSizeInBits.P', 'MaxPSNR',
        'MaxQP', 'MaxQP.B', 'MaxQP.I', 'MaxQP.P', 'MinPSNR', 'MinQP', 'MinQP.B', 'MinQP.I',
        'MinQP.P', 'PBDelta', 'RateCtrlMode', 'ScnChgResilience', 'SCPrevention', 'SliceQP', 'UseGoldenRef'],

        'SETTINGS': ['AspectRatio', 'AvcLowLat', 'BitDepth', 'CabacInit', 'ChromaMode', 'ColourDescription', 'ColourMatrix',
        'Compression', 'ConstrainedIntraPred', 'CostMode', 'CuQpDeltaDepth',
        'DependentSlice', 'DirectMode', 'EnableAUD', 'EnableFillerData', 'EnableSEI', 'EntropyMode', 'FileScalingList', 'LambdaCtrlMode',
        'LambdaFactors', 'Level', 'LookAhead', 'LoopFilter', 'LoopFilter.BetaOffset', 'LoopFilter.CrossSlice', 'LoopFilter.CrossTile',
        'LoopFilter.TcOffset', 'NumCore', 'NumSlices', 'PCM', 'PicCbQpOffset', 'PicCrQpOffset', 'Profile', 'QPCtrlMode',
        'SAO', 'ScalingList', 'SCDFirstPass', 'SliceCbQpOffset', 'SliceCrQpOffset', 'SliceLat', 'SrcFormat', 'StartCode',
        'SubframeLatency', 'Tier', 'TransferCharac', 'TwoPass', 'TwoPassFile', 'UniformSliceType', 'UseL2C', 'VideoFullRange',
        'VideoMode', 'WeightedPred'],

        'RUN': ['BitrateFile', 'FirstPicture',
        'InputSleep', 'Loop', 'MaxPicture', 'ScnChgLookAhead', 'UseBoard']
}

for key, values in keys_values.items():
    param_dict[key] = values

error_dict = {}

error_msg = {
        'Error in ctrlsw app': ['Error', 'error', 'ERROR'],
        'Assertion \'0\' failed': ['Assertion|failed', 'Assertion', 'assertion'],
        'Encoder Resource Error': ['Failed to create Encoder'],
        'CMA allocation Error': ['Cannot allocate memory'],
        'No QP File found': ['No QP File found'],
        'Unknown identifire please check property name': ['unknown identifier'],
        'I/p YUVFile not found': ['Exception caught: Can\'t open file for reading'],
        'Exception caught Error': ['Exception caught'],
        'Get higher Profile to support the usecase': ['getHevcMinimumProfile: Assertion \`0\' failed']
}

for key, values in error_msg.items():
    error_dict[key] = values
stream_4 = {
        "Output/4_streams_1080p60_resolution_1","Output/4_streams_1080p60_resolution_2","Output/4_streams_1080p60_resolution_3","Output/4_streams_1080p60_resolution_4",
        "Output/4_streams_1080p60_resolution_5","Output/4_streams_1080p60_resolution_6","Output/4_streams_1080p60_resolution_7","Output/4_streams_1080p60_resolution_8"
        }
stream_2 = {
        "Output/2_streams_4kp30_resolution_1","Output/2_streams_4kp30_resolution_2","Output/2_streams_4kp30_resolution_3","Output/2_streams_4kp30_resolution_4",
        "Output/2_streams_4kp30_resolution_5","Output/2_streams_4kp30_resolution_6","Output/2_streams_4kp30_resolution_7","Output/2_streams_4kp30_resolution_8",
        "Output/2_streams_4kp30_resolution_9","Output/2_streams_4kp30_resolution_10","Output/2_streams_4kp30_resolution_11","Output/2_streams_4kp30_resolution_12",
        "Output/2_streams_4kp30_resolution_13","Output/2_streams_4kp30_resolution_14","Output/2_streams_4kp30_resolution_15","Output/2_streams_4kp30_resolution_16"
        }

def open_workbook(xls_file, xls_sheet):

    #open workbook
    workbook = openpyxl.load_workbook(xls_file)

    #Select the sheet by Index or name
    workbook_sheet = workbook[xls_sheet]

    return workbook_sheet, workbook

def extract_feature(sheet, row_no):
    #param_values will store the heading of the parameters (e.g Width, Height, Format etc..)
    cell = "A" + str(row_no)
   # print("cell:", cell)
    feature_cell = sheet[cell].value
    feature_cell = "Output/" + str(feature_cell.split(".")[1])
   # print(feature_cell)
    try:
        os.mkdir(feature_cell)
        print("Created ", feature_cell, ": folder")
    except FileExistsError:
        print("folder", feature_cell, "already Exist do you want to continue and replace the data with new one? (y/n)")
        user_input2 = input()
        if user_input2.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return feature_cell

def extract_header(sheet, header_row_number):

    param_values = []
    for cell in sheet[header_row_number]:
        cell_value = cell.value
        param_values.append(cell_value)

    return param_values

def extract_parameters(sheet, next_row, cell_values, output_folder):

    #In XLS we will be using row 4 as the heading of the parameters
    TC_No = sheet.cell(row=next_row, column=1).value
    source_file = 'input_files/input.cfg'
    a = 0
    i = 0
    j = 0


    #lines will store the lines made from parsing the table and that we will insert in the table
    lines = []

    #target_text will be holding the cfg section info that in which section line will be apended.
    target_text = []
    avc_flag = 0

    #This loop generates the lines that needs to be updated on cfg file
    for cell in sheet[next_row]:
        if cell_values[i] == "Profile":
            codec_keyword = "AVC"
            pattern = re.compile(r'{}'.format(codec_keyword))
            #print("!!!!!!!!!!!!!!!!Patern:", pattern)
            match = re.search(pattern, cell.value)
            if match:
               # print("############Match")
                avc_flag = 1
            else:
                hevc_flag = 1
        if cell.value is None:
            if str(cell_values[i]) == "BitstreamFile" and args.output is True:
                if avc_flag == 1:
                    value = str(cell_values[i]) + "      =      " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".avc" + " "
                else:
                    value = str(cell_values[i]) + "      =      " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".hevc" + " "
                    hevc_flag = 0
                target_text.append('OUTPUT')
                lines.append(value)
            if cell_values[i] == "I|YUVFile":
                split_parts = cell_values[i].split("|")
                cell_values[i] = cell_values[i].split("|")[1]
                YUV_Folder = "/mnt/everest/ssw_multimedia_bkup/VCU2/video_YUV/Crowd_Run_" + str(width) + "_" + str(height)
                #print(YUV_Folder)
                search_pattern = f'*_{Format}.*'

                matching_files = glob.glob(f'{YUV_Folder}/{search_pattern}')

                for file_path in matching_files:
                   # print(file_path)
                    value = str(cell_values[i]) + "      =      " + str(file_path) + " "
                    target_text.append('INPUT')
                    lines.append(value)
                cell_values[i] = "|".join(split_parts)
            i = i+1
            continue
        for key, values in param_dict.items():
    #        print("i = ", i)
            if cell_values[i] in values:
     #           print(cell_values[i])
                #This if condition checks for same named parameters and put them in according sections
                if "|" in cell_values[i]:
                    a = 1
                    split_parts = cell_values[i].split("|")
                    cell_values[i] = cell_values[i].split("|")[1]
                    if cell_values[i] == "Width":
                        width = cell.value
                       # print(width)
                    if cell_values[i] == "Height":
                        height = cell.value
                       # print(height)
                    if cell_values[i] == "Format":
                        Format = cell.value
                       # print(Format)
                target_text.append(key)
                break
        value = str(cell_values[i]) + "      =      " + str(cell.value) + " "
        lines.append(value)
        if a == 1:
            cell_values[i] = "|".join(split_parts)
      #      print(cell_values[i])
            a = 0
        i = i+1
    #print(target_text)

    #This block of code generates the cfg files for each testcase
    try:
        os.mkdir(str(output_folder) + "/" + str(TC_No))
        print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
    except FileExistsError:
        print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
        user_input = input()
        if user_input.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()


    destination_file = f'{output_folder}/{TC_No}/input_{TC_No}.cfg'
    shutil.copy2(source_file, destination_file)


    #This block compares the target_text with sections inside cfg and gets the lines no. and it will append the line on next line of
    #matching section
    i = 1
    for k in range(len(lines)-2):
        #print(len(lines))
      #  print(len(target_text))

        with open(destination_file, 'r') as file:
            for line_num, line in enumerate(file, 1):
                if target_text[j] in line:
                    final_line = line_num

            with open(destination_file, 'r') as file:
                    line1 = file.readlines()

            if final_line >= 1 and final_line <= len(line1) + 1:
                    line1.insert(final_line, lines[i] + '\n')

            with open(destination_file, 'w') as file:
                        file.writelines(line1)
            j = j+1
            i = i+1

    return lines

def parce_error(file_path, error_dict):
    with open(file_path, 'r') as file:
        file_contents = file.read()
        for error_message, error_keywords in error_dict.items():
            for keyword in error_keywords:
                if keyword in file_contents:
                    print(f"Error: {error_message}")
                    error = 1
                    return error

def multistream_func(subprocess_pids, test_case_pid, start):

    if (start == 1):

        mapping_dict_test_case = {}

        for element in test_case_pid:
            key, value = element.split('=')
            mapping_dict_test_case[key] = int(value)

        while subprocess_pids:
            pids, status = os.waitpid(0, os.WNOHANG)
            if pids > 0:

                for key, value in mapping_dict_test_case.items():
                    if value == pids:
                        test_case = str(key)
                        break

                print("\n\nCompleted----------------------", test_case, "-----------------------------------\n\n")
                returncode = status >> 8
                print(f"process {pids} completed with return code: {returncode} \n")
                subprocess_pids.remove(pids)
        

def update_xlsx_func(pids_cell_rows, output_string2_pid, output_file_pid, log_file_pid, start, stream_md5_path_pid):    

    if (start == 1):
        print("\n Updating the output.xlsx file... \n")
        mapping_dict_output_string2 = {}
        mapping_dict_output_file = {}
        mapping_dict_log_file = {}
        mapping_dict_stream_md5 = {}

        for element in output_string2_pid:
            value, key = element.split('=')
            mapping_dict_output_string2[int(key)] = value

        for element in output_file_pid:
            value, key = element.split('=')
            mapping_dict_output_file[int(key)] = value

        for element in log_file_pid:
            value, key = element.split('=')
            mapping_dict_log_file[int(key)] = value

        for element in stream_md5_path_pid:
            value, key = element.split('=')
            mapping_dict_stream_md5[int(key)] = value

        for row_index in pids_cell_rows:
#            print(f"\n row_index : '{row_index}'")
            output_string2 = mapping_dict_output_string2[row_index]
            output_string2 = output_string2.strip()
 #           print(f"\n output_string2 : '{output_string2}'")
            output_file_path = mapping_dict_output_file[row_index]
            output_file_path = output_file_path.strip()
#            print(f"\n out_file_path : '{output_file_path}'")
            log_file = mapping_dict_log_file[row_index]
            log_file = log_file.strip()
#            print(f"\n log_file : '{log_file}'")
            stream_md5_path = mapping_dict_stream_md5[row_index]
            stream_md5_path = stream_md5_path.strip()
            with open(stream_md5_path, 'r', encoding='utf-8') as file:
                stream_md5_contents = file.read()
                stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
                stream_md5_contents = stream_md5_contents.split(" ")[0]
            md5_command = ['md5sum',output_file_path]
            md5_process = subprocess.Popen(md5_command, stdout=subprocess.PIPE)
            output, error = md5_process.communicate()
            output_file = output_file_path.split("/")[-1]
            if md5_process.returncode == 0:
                hw_md5_content = output.decode().split()[0]
            else:
                hw_md5_content = " "

            if hw_md5_content == stream_md5_contents:
                md5_flag = 1
            else:
                md5_flag = 0
            if yuv_result_flag != 1:
           #     print("row_index : ", row_index, "yuv_index+1 : ", (yuv_index+1))
                yuv_result_cell = output_sheet.cell(row=int(row_index),column=(yuv_index+1))
                yuv_result_cell.value = output_string2
            if b_result_flag != 1:
                b_result_cell = output_sheet.cell(row=int(row_index),column=(b_index+1))
                b_result_cell.value = output_file
            if stream_md5_flag != 1:
                stream_md5_cell = output_sheet.cell(row=int(row_index),column=(stream_md5_index+1))
                stream_md5_cell.value = stream_md5_contents
            if hw_md5_flag != 1:
                hw_md5_cell = output_sheet.cell(row=int(row_index),column=(hw_md5_index+1))
                hw_md5_cell.value = hw_md5_content
            error_flag = parce_error(log_file, error_dict)

            if error_flag != 1 and result_flag != 1 and md5_flag == 1:
                result_cell = output_sheet.cell(row=int(row_index),column=(index+1))
                result_cell.value = "PASS"
            else:
                result_cell = output_sheet.cell(row=int(row_index),column=(index+1))
                result_cell.value = "FAIL"
            output_workbook.save(output_xls)

        print("\n output.xlsx file Updated \n")


#----------------------------------------------------------------------------------------------------------------------#

parser = argparse.ArgumentParser(description='Testcase_automation_V1.0', add_help=True)

parser.add_argument('-f', '--file', help='Specify a input XLS file for testcases')
parser.add_argument('-s', '--sheet', help='Specify a input XLS file\'s sheet for testcases')
parser.add_argument('-o', '--output', action='store_true', default=False, help='Select this option to store the output file')
parser.add_argument('-tc', '--tc_no', help='Specify individual TC_No that u want to run e.g -tc TC_0001')

args = parser.parse_args()

file_option = args.file
sheet_option = args.sheet
output_option = args.output

#print(f"file_option: {file_option}")
#print(f"sheet_option: {sheet_option}")
#print(f"output_option: {output_option}")

subprocess_pids = []
pids_cell_rows = []
output_file_pid = []
output_string2_pid = []
log_file_pid = []
test_case_pid = []
stream_md5_path_pid = []

try:
    os.mkdir("Output")
    print("Created Output folder: Head over to this folder for more TC related information and output files")
except FileExistsError:
    user_input1 = input("Output folder alreaady Exist do you want to continue and replace the data with new one? (y/n)")
    if user_input1.lower() == 'y':
        pass
    else:
        print("Program closing")
        exit()
#xls_file = '/group/siv3/staff/andreis/sibridge/yashl/VCU2/python_script_for_TC/yash_python_scripting/XLS/Encoder_tests.xlsx'
#xls_sheet = 'Temp'
#next_row = 4

CWD = os.getcwd()
#print(CWD)
call_multistream = 0
orignal_xls = args.file
output_xls = "Output/output.xlsx"

shutil.copy2(orignal_xls, output_xls)

output_sheet, output_workbook = open_workbook(str(output_xls), str(args.sheet))

sheet, new_workbook = open_workbook(str(args.file), str(args.sheet))
#next_row, header_values, Heading_cell = extract_headers(sheet)

for cell in sheet['A']:
    fill_color = cell.fill.start_color.rgb
   # print(f"The fill color of the cell is: {fill_color}")
    if sheet[cell.coordinate].fill.start_color.rgb == 'FFFF0000':
        break
    if sheet[cell.coordinate].fill.start_color.rgb != 'FF000000' and cell.value is None:
    #    print("Continue", cell.row)
        continue
    if sheet[cell.coordinate].fill.start_color.rgb == 'FF000000':
        #Enabling flag so in next row we will extract feature name
        multistream_func(subprocess_pids,test_case_pid, call_multistream)
        update_xlsx_func(pids_cell_rows, output_string2_pid, output_file_pid, log_file_pid, call_multistream, stream_md5_path_pid)
        subprocess_pids.clear()
        pids_cell_rows.clear()
        output_file_pid.clear()
        output_string2_pid.clear()
        log_file_pid.clear()
        test_case_pid.clear()
        stream_md5_path_pid.clear()
        call_multistream = 1
        extract_feature_flag = 1
     #   print("Black detected at row:", cell.row)
        continue
    if extract_feature_flag == 1:
      #  print("extract_feature condition true")
        feature_folder = extract_feature(sheet, cell.row)
        log_folder = str(CWD) + "/" + str(feature_folder)
        if feature_folder == "Output/32_streams_mix_resolution":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_32/32_streams_mix_resolution"
        elif feature_folder == "Output/32_streams_480p30_resolution":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_32/32_streams_480p30_resolution"
        elif feature_folder == "Output/8_streams_1080p30_resolution_1":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_8/8_streams_1080p30_resolution_1"
        elif feature_folder == "Output/8_streams_1080p30_resolution_2":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_8/8_streams_1080p30_resolution_2"
        elif feature_folder == "Output/8_streams_1080p30_resolution_3":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_8/8_streams_1080p30_resolution_3"
        elif feature_folder == "Output/8_streams_1080p30_resolution_4":
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_8/8_streams_1080p30_resolution_4"
        elif feature_folder in stream_4:
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_4"
        elif feature_folder in stream_2: 
            stream_md5_path = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Multistream/Output_2"
        else:
            print("Error: Unexpected featute folder name")
       # print(log_folder)
        extract_feature_flag = 0
        #We got the feature name enabling this flag as in next row we will extract the headers of testcase
        extract_header_flag = 1
        continue
    if extract_header_flag == 1:
       # print("Extract headers condition true")
        header_values = extract_header(sheet, cell.row)
       # print("###############", "row", cell.row, len(header_values))
        extract_header_flag = 0
        continue
    if args.tc_no is not None:
#        print("Tc argumenat found")
        if cell.value != args.tc_no:
            continue
    parameters = extract_parameters(sheet, cell.row, header_values, log_folder)
    test_case = str(parameters[0].split("=")[1])
    test_case = test_case.replace(" ","")
    md5_file = test_case + ".md5"
    print("Running----------------------", test_case, "-----------------------------------\n\n")
    if "=" in parameters[0]:
        log_file = log_folder + "/" + cell.value + "/" + str(parameters[0].split("=")[1]) + ".txt"
        log_file = log_file.replace(" ","")
   #     print("log file:", log_file)
    with open(log_file, "w") as file:
        current_time = datetime.datetime.now()
        #this the maximum time we will wait for 1 usecase
        deadline = current_time + datetime.timedelta(minutes=5)
        command = "ctrlsw_encoder --embedded --device /dev/al_e2xx -cfg " + str(log_folder) + "/" + str(test_case) + "/" + "input_" + str(test_case) + ".cfg"
        print(command)
        #process = subprocess.Popen(command, shell=True, stdout=file, text=True)
        process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        pid = process.pid
        subprocess_pids.append(pid)
       # print("/-/-/-/-/\n", subprocess_pids)
       # time.sleep(3)
        if process.poll() is None:
            pass
        else:
        #    print("Remooooooooooooooooooooving pid  :  ", pid)
            subprocess_pids.remove(pid)
        #print("After removing /-/-/-/-/\n", subprocess_pids)
        #Polling here until the encoding or decoding is Done
#        while process.poll() is None:
#            time.sleep(5)
#            if datetime.datetime.now() > deadline:
#                time_failure = 1
#                break
    print("\n\n", parameters, "\n\n")
    substring = "BitstreamFile"
    filtered_list = [element for element in parameters if substring in element]
   # print(filtered_list)
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
       # print("@!@!@!@!@!@!@!@!",final_index)
        output_string = parameters[final_index].split("=")[1]
        output_file = output_string
       # output_file = output_string.split("/")[-1]

       # print(output_file)
    substring2 = "YUVFile"
    filtered_list2 = [element for element in parameters if substring2 in element]
   # print(filtered_list)
    if filtered_list2:
        final_index2 = parameters.index(filtered_list2[0])
       # print("@!@!@!@!@!@!@!@!",final_index)
        output_string2 = parameters[final_index2].split("=")[1]
       # print(output_string2)
       # print(output_file)
    try:
        yuv_index = header_values.index('I|YUVFile')
        yuv_result_flag = 0
    except:
        yuv_result_flag = 1
    try:
        b_index = header_values.index('BitstreamFile')
        b_result_flag = 0
    except:
        b_result_flag = 1
    try:
        stream_md5_index = header_values.index('Stream MD5sum')
        stream_md5_flag = 0
    except:
        stream_md5_flag = 1
    try:
        hw_md5_index = header_values.index('HW MD5sum')
        hw_md5_flag = 0
    except:
        hw_md5_flag = 1
    try:
        index = header_values.index('Result')
        result_flag = 0
    except:
        result_flag = 1
  #  print("!!!!!!!!!!!!!!!   Index:", index)
 #   print(len(header_values))
    log_file_pid_value = str(log_file) + "=" + str(cell.row)
    #print(log_file_pid_value)
    log_file_pid.append(log_file_pid_value)
    pid_cell_row_value = cell.row
    #print(pid_cell_row_value)
    output_string2_pid_value = str(output_string2) + "=" + str(cell.row)
    output_string2_pid.append(output_string2_pid_value)
    #print(output_string2_pid_value)
    test_case_pid_value = str(test_case) + "=" + str(pid)
    test_case_pid.append(test_case_pid_value)
    output_file_pid_value = str(output_file) + "=" + str(cell.row)
    output_file_pid.append(output_file_pid_value)
    #print(output_file_pid_value)
    for root, dirs, files in os.walk(stream_md5_path):
        if md5_file in files:
            stream_md5_file = os.path.join(root,md5_file)
    stream_md5_path_value = str(stream_md5_file) + "=" + str(cell.row)
    stream_md5_path_pid.append(stream_md5_path_value)
    pids_cell_rows.append(pid_cell_row_value)


