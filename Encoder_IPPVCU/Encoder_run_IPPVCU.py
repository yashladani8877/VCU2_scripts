import os
import io
import sys
import openpyxl
import shutil
import fileinput
import subprocess
import time
import datetime
import argparse
import re
import glob
import signal

param_dict = {}
keys_values = {
        'GOP': ['Gop.DoubleRef', 'Gop.EnableLT', 'Gop.FreqIDR', 'Gop.FreqLT', 'Gop.FreqRP', 'Gop.GdrMode', 'Gop.Length',
        'Gop.NumB', 'Gop.TempDQP', 'Gop.WriteAVCHeaderSVC', 'GopCtrlMode'],

        'INPUT': ['CmdFile', 'I|CropHeight', 'I|CropPosX', 'I|CropPosY', 'I|CropWidth', 'I|Format', 'I|FrameRate', 'GMVFile', 'HDRFile',
        'I|Height', 'MapFile', 'QpTablesFolder', 'ROIFile', 'I|Width', 'I|YUVFile'],

        'DYNAMIC_INPUT': ['D|Height1', 'D|Width1', 'D|YUVFile1', 'D|Height2', 'D|Width2', 'D|YUVFile2', 'D|Height3', 'D|Width3', 'D|YUVFile3'],

        'OUTPUT': ['BitstreamFile', 'O|CropHeight', 'O|CropPosX', 'O|CropPosY', 'O|CropWidth', 'O|Format', 'RecFile'],

        'RATE_CONTROL': ['BitRate', 'CPBSize', 'EnableSkip', 'FrameRate', 'InitialDelay', 'IPDelta', 'MaxBitRate',
        'MaxConsecutiveSkip', 'MaxPictureSize', 'MaxPictureSize.B', 'MaxPictureSize.I', 'MaxPictureSize.P',
        'MaxPictureSizeInBits', 'MaxPictureSizeInBits.B', 'MaxPictureSizeInBits.I', 'MaxPictureSizeInBits.P', 'MaxPSNR',
        'MaxQP', 'MaxQP.B', 'MaxQP.I', 'MaxQP.P', 'MinPSNR', 'MinQP', 'MinQP.B', 'MinQP.I',
        'MinQP.P', 'PBDelta', 'RateCtrlMode', 'ScnChgResilience', 'SCPrevention', 'SliceQP', 'UseGoldenRef'],

        'SETTINGS': ['AspectRatio', 'AvcLowLat', 'BitDepth', 'CabacInit', 'ChromaMode', 'ColourDescription', 'ColourMatrix',
        'Compression', 'ConstrainedIntraPred', 'CostMode', 'CuQpDeltaDepth',
        'DependentSlice', 'DirectMode', 'EnableAUD', 'EnableFillerData', 'EnableSEI', 'EntropyMode', 'FileScalingList', 'LambdaCtrlMode',
        'LambdaFactors', 'Level', 'LookAhead', 'LoopFilter', 'LoopFilter.BetaOffset', 'LoopFilter.CrossSlice', 'LoopFilter.CrossTile',
        'LoopFilter.TcOffset', 'NumCore', 'NumSlices', 'PCM', 'PicCbQpOffset', 'PicCrQpOffset', 'Profile', 'QPCtrlMode',
        'SAO', 'ScalingList', 'SCDFirstPass', 'SliceCbQpOffset', 'SliceCrQpOffset', 'SliceLat', 'SrcFormat', 'StartCode',
        'SubframeLatency', 'Tier', 'TransferCharac', 'TwoPass', 'TwoPassFile', 'UniformSliceType', 'UseL2C', 'VideoFullRange',
        'VideoMode', 'WeightedPred'],

        'RUN': ['BitrateFile', 'FirstPicture','First Picture',
        'InputSleep', 'Loop', 'MaxPicture', 'ScnChgLookAhead', 'UseBoard']
}

# iterates through the key-value pairs, and it adds the entry to another dictionary `param_dict`.

for key, values in keys_values.items():
    param_dict[key] = values

# This dictionary contains predefined error messages,                                                                                             
# key : represents a general description of the error, and 
# value : is a list of specific strings that are used to identify these errors in logs or output.

error_dict = {}
error_msg = {
        'Error in ctrlsw app': ['Error', 'error', 'ERROR'],
        'Assertion \'0\' failed': ['Assertion|failed', 'Assertion', 'assertion'],
        'Encoder Resource Error': ['Failed to create Encoder'],
        'CMA allocation Error': ['Cannot allocate memory'],
        'No QP File found': ['No QP File found'],
        'Unknown identifire please check property name': ['unknown identifier'],
        'I/p YUVFile not found': ['Exception caught: Can\'t open file for reading'],
        'Exception caught Error': ['Exception caught'],
        'Get higher Profile to support the usecase': ['getHevcMinimumProfile: Assertion \'0\' failed']
}

# iterates through the key-value pairs, and it adds the entry to another dictionary `error_dict`.

for key, values in error_msg.items():
    error_dict[key] = values

# Open an Excel file and Select a specific sheet from that file.
# Return both the selected sheet and the whole workbook for further manipulation.

def open_workbook(xls_file, xls_sheet):
    workbook = openpyxl.load_workbook(xls_file)
    workbook_sheet = workbook[xls_sheet]
    return workbook_sheet, workbook

# create a folder, checks if the folder already exists and It then returns the folder path for further use.

def extract_feature(sheet, row_no):
    cell = "A" + str(row_no)
    feature_cell = sheet[cell].value
    feature_cell = "Output/" + str(feature_cell.split(".")[1])
   
    try:
        os.mkdir(feature_cell)
        print("Created ", feature_cell, ": folder")
    except FileExistsError:
        print("folder", feature_cell, "already Exist do you want to continue and replace the data with new one? (y/n)")
        user_input2 = input()
        if user_input2.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return feature_cell

# reads the header row of an Excel sheet, collects the values from each cell in that row, and returns them as a list.

def extract_header(sheet, header_row_number):
    # param_values will store the heading of the parameters (e.g Width, Height, Format etc..) 
    param_values = []
    for cell in sheet[header_row_number]:
        cell_value = cell.value
        param_values.append(cell_value)

    return param_values

# function extracts parameter values from an Excel sheet, processes YUV file paths, and 
# skips empty or invalid cells, ensuring valid inputs for further script execution.
# return processed list of parameter values

def extract_parameters_omx(sheet_omx, next_row_omx, cell_values_omx, output_folder_omx):

    #This function will be called in case of OpenMax TC only
    TC_No = sheet_omx.cell(row=next_row_omx, column=1).value
    a = 0
    i = 0
    j = 0
    skip = 0

    lines_omx = []

    for cell in sheet_omx[next_row_omx]:
        if cell_values_omx[i] == "Result" or cell_values_omx[i] == "TC_No" or cell_values_omx[i] == "HD MD5sum" or cell_values_omx[i] == "Comment":
            if cell.value == "PASS":
                skip = 1
                break
            i = i + 1
            continue

        if cell_values_omx[i] == "codec":
            codec = cell.value
            value = "--" + str(cell.value)
            lines_omx.append(value)
            i = i + 1
            continue

        #Non breaking space replace with None
        if cell.value == '\xa0':
            cell.value = None

        if cell.value == None and cell_values_omx[i] != "input_file" and cell_values_omx[i] != "out":
            i = i + 1
            continue

        if cell_values_omx[i] == "width":
            width_omx = cell.value
        elif cell_values_omx[i] == "height":
            height_omx = cell.value
        elif cell_values_omx[i] == "fourcc":
            fourcc_omx = cell.value
        elif cell_values_omx[i] == "dma-out" or cell_values_omx[i] == "dma-in":
            if cell.value == 0:
                i = i + 1
                continue
            else:
                lines_omx.append("--" + str(cell_values_omx[i]))
                i = i + 1
                continue


        #Add logic to find the input YUV file from width,height and format
        if cell_values_omx[i] == "out":
            if 'hevc' in str(codec): 
                out_file = str(log_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".hevc"
            else:
                out_file = str(log_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".avc"
            value = "--" + str(cell_values_omx[i]) + " " + str(out_file)
            lines_omx.append(value)
            i = i + 1
            continue
        if cell_values_omx[i] == "input_file":
            YUV_Folder_omx = "/mnt/build/ssw_vcu/yashl/VCU2/video_YUV/Crowd_Run_" + str(width_omx) + "_" + str(height_omx)
            search_pattern_omx = f'*_{fourcc_omx}.*'
            #search_pattern_omx = search_pattern_omx.upper()

            matching_files_omx = glob.glob(f'{YUV_Folder_omx}/{search_pattern_omx}')


            for file_path in matching_files_omx:
                value = str(file_path)
                lines_omx.append(value)

            i = i + 1
            continue
        if cell_values_omx[i] == "fourcc":
            value = "--" + str(cell_values_omx[i]) + " " + str(cell.value.lower())
        else:
            value = "--" + str(cell_values_omx[i]) + " " + str(cell.value)
        lines_omx.append(value)
        i = i + 1
        continue
    try:
        os.mkdir(str(output_folder_omx) + "/" + str(TC_No))
        print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
    except FileExistsError:
        print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
        user_input = input()
        if user_input.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return lines_omx



def extract_parameters(sheet, next_row, cell_values, output_folder):

    #In XLS we will be using row 4 as the heading of the parameters
    TC_No = sheet.cell(row=next_row, column=1).value
    source_file = 'input_files/input.cfg'
    a = 0
    i = 0
    j = 0
    skip = 0


    #lines will store the lines made from parsing the table and that we will insert in the table
    lines = []

    #target_text will be holding the cfg section info that in which section line will be apended.
    target_text = []
    avc_flag = 0

    #This loop generates the lines that needs to be updated on cfg file
    for cell in sheet[next_row]:
        if cell_values[i] == "Result":
            if cell.value == "PASS":
                skip = 1
            break
        if cell_values[i] == "Profile":
            codec_keyword = "AVC"
            pattern = re.compile(r'{}'.format(codec_keyword))
            
            match = re.search(pattern, cell.value)
            if match:
                avc_flag = 1
            else:
                hevc_flag = 1

        #Non breaking space replace with None
        if cell.value == '\xa0':
            cell.value = None 

        if cell.value is None: 
            if str(cell_values[i]) == "BitstreamFile" and args.output is True:
                if avc_flag == 1:
                    value = str(cell_values[i]) + "      =      " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".avc" + " "
                else:
                    value = str(cell_values[i]) + "      =      " + str(output_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".hevc" + " "
                    hevc_flag = 0
                target_text.append('OUTPUT')
                lines.append(value)

            if cell_values[i] == "I|YUVFile":
                split_parts = cell_values[i].split("|")
                cell_values[i] = cell_values[i].split("|")[1]
                YUV_Folder = "/mnt/build/ssw_vcu/yashl/VCU2/video_YUV/Crowd_Run_" + str(width) + "_" + str(height)
                search_pattern = f'*_{Format}.*'

                matching_files = glob.glob(f'{YUV_Folder}/{search_pattern}')

                for file_path in matching_files:
                    value = str(cell_values[i]) + "      =      " + str(file_path) + " "
                    target_text.append('INPUT')
                    lines.append(value)
                cell_values[i] = "|".join(split_parts)

            if cell_values[i] == "D|YUVFile1":
                split_parts = cell_values[i].split("|")
                cell_values[i] = cell_values[i].split("|")[1]
                YUV_Folder = "/mnt/build/ssw_vcu/yashl/VCU2/video_YUV/Crowd_Run_" + str(width1) + "_" + str(height1)
                search_pattern = f'*_{Format}.*'

                matching_files = glob.glob(f'{YUV_Folder}/{search_pattern}')

                for file_path in matching_files:
                    value = str(cell_values[i]) + "      =      " + str(file_path) + " "
                    target_text.append('DYNAMIC_INPUT')
                    lines.append(value)
                cell_values[i] = "|".join(split_parts)

            if cell_values[i] == "D|YUVFile2":
                split_parts = cell_values[i].split("|")
                cell_values[i] = cell_values[i].split("|")[1]
                YUV_Folder = "/mnt/build/ssw_vcu/yashl/VCU2/video_YUV/Crowd_Run_" + str(width2) + "_" + str(height2)
                search_pattern = f'*_{Format}.*'

                matching_files = glob.glob(f'{YUV_Folder}/{search_pattern}')

                for file_path in matching_files:
                    value = str(cell_values[i]) + "      =      " + str(file_path) + " "
                    target_text.append('DYNAMIC_INPUT')
                    lines.append(value)
                cell_values[i] = "|".join(split_parts)

            if cell_values[i] == "D|YUVFile3":
                split_parts = cell_values[i].split("|")
                cell_values[i] = cell_values[i].split("|")[1]
                YUV_Folder = "/mnt/build/ssw_vcu/yashl/VCU2/video_YUV/Crowd_Run_" + str(width3) + "_" + str(height3)
                search_pattern = f'*_{Format}.*'

                matching_files = glob.glob(f'{YUV_Folder}/{search_pattern}')

                for file_path in matching_files:
                    value = str(cell_values[i]) + "      =      " + str(file_path) + " "
                    target_text.append('DYNAMIC_INPUT')
                    lines.append(value)
                cell_values[i] = "|".join(split_parts)

            i = i+1
            continue

        for key, values in param_dict.items():
            if cell_values[i] in values:

                #This if condition checks for same named parameters and put them in according sections
                if "|" in cell_values[i]:
                    a = 1
                    split_parts = cell_values[i].split("|")
                    cell_values[i] = cell_values[i].split("|")[1]
                    if cell_values[i] == "Width":
                        width = cell.value
                    if cell_values[i] == "Height":
                        height = cell.value
                    if cell_values[i] == "Width1":
                        width1 = cell.value
                    if cell_values[i] == "Height1":
                        height1 = cell.value
                    if cell_values[i] == "Width2":
                        width2 = cell.value
                    if cell_values[i] == "Height2":
                        height2 = cell.value
                    if cell_values[i] == "Width3":
                        width3 = cell.value
                    if cell_values[i] == "Height3":
                        height3 = cell.value
                    if cell_values[i] == "Format":
                        Format = cell.value
                target_text.append(key)
                break
        value = str(cell_values[i]) + "      =      " + str(cell.value) + " "
        lines.append(value)
        if a == 1:
            cell_values[i] = "|".join(split_parts)
            a = 0
        i = i+1


    #This block of code generates the cfg files for each testcase
    if skip != 1:
        try:
            os.mkdir(str(output_folder) + "/" + str(TC_No))
            print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
        except FileExistsError:
            print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
            user_input = input()
            if user_input.lower() == 'y':
                pass
            else:
                print("Program closing")
                exit()
    if skip != 1:
        destination_file = f'{output_folder}/{TC_No}/input_{TC_No}.cfg'
        shutil.copy2(source_file, destination_file)


    #This block compares the target_text with sections inside cfg and gets the lines no. and it will append the line on next line of
    #matching section
    i = 1
    for k in range(len(lines)-1):
        execute_once = True
    
        if skip != 1:
            with open(destination_file, 'r') as file:
                try : 
                    for line_num, line in enumerate(file, 1):
                        if target_text[j] in line:
                            final_line = line_num

                            if target_text[j] == "DYNAMIC_INPUT":
                                match = re.search(r'([a-zA-Z]+)(\d*)\s*=', lines[i])
                                if match:
                                    # Extract the captured groups
                                    alpha_part = match.group(1)
                                    trailing_digits = match.group(2)

                                modified_string = re.sub(r'([a-zA-Z]+)\d*\s*=', r'\1 =', lines[i])
             
                                if trailing_digits == "2" and execute_once:
                                    execute_once = False
                                    continue
                                if trailing_digits == "3":
                                    continue
                            execute_once = True
                            break
                except IndexError as e :
                    print(f"Error : {e} \n* Please make sure Bitstreamfile, Stream Md5sum, HW Md5sum and Result fields in xlsx file are empty *")

                with open(destination_file, 'r') as file:
                        line1 = file.readlines()

                if final_line >= 1 and final_line <= len(line1) + 1:
                    if target_text[j] != "DYNAMIC_INPUT":
                        line1.insert(final_line, lines[i] + '\n')
                    else:
                        line1.insert(final_line, modified_string + '\n')

                with open(destination_file, 'w') as file:
                            file.writelines(line1)
                j = j+1
                i = i+1
    return lines

# Identify known issues or failures by checking logs predefined error message

def parce_error(file_path, error_dict):
    with open(file_path, 'r') as file:
        file_contents = file.read()
        for error_message, error_keywords in error_dict.items():
            for keyword in error_keywords:
                if keyword in file_contents:
                    print(f"Error: {error_message}")
                    error = 1
                    return error



#----------------------------------------------------------------------------------------------------------------------#

parser = argparse.ArgumentParser(description='Testcase_automation_V1.0', add_help=True)

parser.add_argument('-f', '--file', help='Specify a input XLS file for testcases')
parser.add_argument('-s', '--sheet', help='Specify a input XLS file\'s sheet for testcases')
parser.add_argument('-o', '--output', action='store_true', default=False, help='Select this option to store the output file')
parser.add_argument('-tc', '--tc_no', help='Specify individual TC_No that u want to run e.g -tc TC_0001')

args = parser.parse_args()

file_option = args.file
sheet_option = args.sheet
output_option = args.output

# create a folder, check if the folder already exists,
# and prompt the user whether to overwrite it or not.

try:
    os.mkdir("Output")
    print("Created Output folder: Head over to this folder for more TC related information and output files")
except FileExistsError:
    user_input1 = input("Output folder alreaady Exist do you want to continue and replace the data with new one? (y/n)")
    if user_input1.lower() == 'y':
        pass
    else:
        print("Program closing")
        exit()

CWD = os.getcwd()
omx_flag = 0

orignal_xls = args.file
output_xls = "Output/output.xlsx"

shutil.copy2(orignal_xls, output_xls)

output_sheet, output_workbook = open_workbook(str(output_xls), str(args.sheet))

sheet, new_workbook = open_workbook(str(args.file), str(args.sheet))

for cell in sheet['A']:
    time_failure = 0
    if sheet[cell.coordinate].fill.start_color.rgb == 'FFFF0000': #FF0000 is Red
        break
    if sheet[cell.coordinate].fill.start_color.rgb != 'FF000000' and cell.value is None:
        continue
    if sheet[cell.coordinate].fill.start_color.rgb == 'FF000000': #000000 is black
        #Enabling flag so in next row we will extract feature name
        extract_feature_flag = 1
        continue
    if extract_feature_flag == 1:
        feature_folder = extract_feature(sheet, cell.row)
        log_folder = str(CWD) + "/" + str(feature_folder)
        if feature_folder == "Output/Color_Format":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Color_Format/Output/Color_Format" 
        elif feature_folder == "Output/Conformance":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Conformance/Output/Conformance"
        elif feature_folder == "Output/GOP":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/GOP"
        elif feature_folder == "Output/Input":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/Input"
        elif feature_folder == "Output/Dynamic_Input":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/Dynamic_Input"
        elif feature_folder == "Output/Settings":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/Settings"
        elif feature_folder == "Output/Output":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/Output"
        elif feature_folder == "Output/Run":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/Run"
        elif feature_folder == "Output/RateControl":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Parameters/Output/RateControl"
        elif feature_folder == "Output/Dynamic_Bframes":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_Bframes"
        elif feature_folder == "Output/Dynamic_Bitrate":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_Bitrate"
        elif feature_folder == "Output/Dynamic_FrameRate":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_FrameRate"
        elif feature_folder == "Output/Dynamic_GOP":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_GOP"
        elif feature_folder == "Output/Dynamic_KeyFrame":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_KeyFrame"
        elif feature_folder == "Output/Dynamic_KFandGOP":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Dynamic_Parameters/Output/Dynamic_KFandGOP"
        elif feature_folder == "Output/OpenMax":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/omx_support/Output/OpenMax" 
            omx_flag = 1
        elif feature_folder == "Output/Performance_filesrc":
            stream_md5_path = "/mnt/build/ssw_vcu/yashl/VCU2/regression_logs/Encoder/Performance/Output/Performance_filesrc"
        else:
            print("Error: Unexpected featute folder name")

        extract_feature_flag = 0
        #We got the feature name enabling this flag as in next row we will extract the headers of testcase
        extract_header_flag = 1
        continue
    if extract_header_flag == 1:
        header_values = extract_header(sheet, cell.row)
        extract_header_flag = 0
        continue
    if args.tc_no is not None:
        if cell.value != args.tc_no:
            continue
    if omx_flag == 1:
        parameters = extract_parameters_omx(sheet, cell.row, header_values, log_folder)
        yuv_index_omx = [i for i, j in enumerate(parameters) if 'yuv' in j]
        yuv_index_omx = yuv_index_omx[0]

        if 'input_file' in header_values:
            input_index_omx = header_values.index('input_file')
            input_file_omx = parameters[yuv_index_omx]
        if 'out' in header_values:
            out_index_omx = header_values.index('out')
    else:
        parameters = extract_parameters(sheet, cell.row, header_values, log_folder)
    substring = "Result"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
        output_string = parameters[final_index].split("=")[1]
        output_string = output_string.replace(" ","")
        if output_string == "PASS":
            continue
    bitstream_substring = "BitstreamFile"
    bitstream_filtered_list = [element for element in parameters if bitstream_substring in element]
    if bitstream_filtered_list:
        temp_index = parameters.index(bitstream_filtered_list[0])
        bitstream_file = parameters[temp_index].split("=")[1]
        bitstream_file = bitstream_file.replace(" ","")

    
    test_case = cell.value
    print("Running----------------------", test_case, "-----------------------------------\n\n")
    log_file = log_folder + "/" + cell.value + "/" + str(cell.value) + ".txt"
    log_file = log_file.replace(" ","")
    md5_file = log_file.split(".")[0] + ".md5"

    try :
        with open(log_file, "w") as file:
            current_time = datetime.datetime.now()
            mem_command = "cat /proc/meminfo"
            process = subprocess.Popen(mem_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
            process.wait()
            #this the maximum time we will wait for 1 usecase
            deadline = current_time + datetime.timedelta(minutes=180)
            if sheet_option == 'Performance':
                performance_enc_flag = 1
                # Extracting required parameter to create gstreamer pipeline
                for param in parameters:
                    if 'MaxPicture' in param:
                        maxpicture = param.split('=')[-1].strip()
                    if 'FrameRate' in param :
                        framerate = param.split('=')[-1].strip()
                    if 'Width' in param :
                        width =  param.split('=')[-1].strip()
                    if 'Height' in param :
                        height = param.split('=')[-1].strip()
                    if 'Format' in param:
                        format = param.split('=')[-1].strip()
                        if format == 'I4CL':
                            format = 'y444-12le'
                        if format == 'NV12':
                            format = 'nv12'
                        if format == 'P012':
                            format = 'p012-le'
                        if format == 'P212':
                            format = 'p212-12le'
                        if format == 'T50C':
                            format = 't50c'
                        if format == 'T52C':
                            format = 't52c'
                        if format == 'T54C':
                            format = 't54c'
                        if format == 'T5MC':
                            format = 't5mc'
                        if format == 'T60C':
                            format = 't60c'
                        if format == 'T62C':
                            format = 't62c'
                        if format == 'T64C':
                            format = 't64c'
                        if format == 'T6MC':
                            format = 't6mc'
                        if format == 'Y012':
                             format = 'gray12-le'
                    if 'YUVFile' in param:
                        yuvFile = param.split('=')[-1].strip()
                    if 'BitstreamFile' in param:
                        encoded_file = param.split('=')[-1].strip()
                        if encoded_file.endswith('.avc'):
                            omx_enc = 'omxh264enc'
                        elif encoded_file.endswith('.hevc'):
                            omx_enc = 'omxh265enc'

                command =  f"gst-launch-1.0 filesrc location={yuvFile} num-buffers={maxpicture} ! rawvideoparse format={format}  width={width} height={height} framerate={framerate}/1 ! {omx_enc} !  filesink location={encoded_file}"
            elif omx_flag == 1:
                command = "omx_encoder " + str(parameters[yuv_index_omx])
                for x in parameters:
                    if 'yuv' in x:
                        continue
                    elif 'out' in x:
                        bitstream_omx = str(x)
                        bitstream_omx = bitstream_omx.split(" ")[1]
                        command = command + " " + str(x)
                    else:
                        command = command + " " + str(x)
            else:    
                command = "ctrlsw_encoder --embedded --device /dev/al_e2xx -cfg " + str(log_folder) + "/" + str(test_case) + "/" + "input_" + str(test_case) + ".cfg " + "--md5-stream " + str(md5_file)

            print(command)
        
            process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
            pid = process.pid
            #Polling here until the encoding or decoding is Done
            while process.poll() is None:
                time.sleep(5)
                if datetime.datetime.now() > deadline:
                    time_failure = 1
                    try:
                        os.kill(pid, signal.SIGTERM)
                        print("Due to hang, ", test_case, "is failed and process : ", pid, "is killed")
                        break
                    except OSError:
                        print("Could not kill the process : ", pid)
#                       break
    
            process = subprocess.Popen(mem_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
            process.wait()
    except FileNotFoundError as e:
        print(f"Error : {e}, The file path does not exist.\n* Please check input YUV file *")

    print("\n\n", parameters, "\n\n")
    if omx_flag == 1:
        bitstream_file = bitstream_omx

    md5_command = ['md5sum', bitstream_file]
    md5_process = subprocess.Popen(md5_command, stdout=subprocess.PIPE)
    output, error = md5_process.communicate()
    if md5_process.returncode == 0:
        hw_md5_content = output.decode().split()[0]
    else:
        print(f'md5 Error: {md5_command}')
        hw_md5_content = " "
    if performance_enc_flag == 1:
        with open (md5_file, 'w') as file:
            file.writelines(hw_md5_content)
    if omx_flag == 1:
        with open (md5_file, 'w') as file:
            file.writelines(hw_md5_content)
   
    stream_md5_file = stream_md5_path + "/" + test_case + "/" + test_case + ".md5"

    if os.path.exists(stream_md5_file):
        with open(stream_md5_file, 'r', encoding='utf-8') as file:
            stream_md5_contents = file.read()
            stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
            stream_md5_contents = stream_md5_contents.split(" ")[0]
    else:
        print("Error reading",stream_md5_file)
        stream_md5_contents = " "
    if hw_md5_content == stream_md5_contents :
        md5_flag = 1
    else:
        md5_flag = 0

    substring = "BitstreamFile"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
        output_string = parameters[final_index].split("=")[1]
        output_file = output_string.split("/")[-1]
    
    if omx_flag == 1:
        substring = "out"
        filtered_list = [element for element in parameters if substring in element]
        if filtered_list:
            final_index_omx = parameters.index(filtered_list[0])
            output_string_omx = parameters[final_index_omx].split(" ")[1]
            output_file_omx = output_string_omx.split("/")[-1]


    substring2 = "YUVFile"
    filtered_list2 = [element for element in parameters if substring2 in element]
    if filtered_list2:
        final_index2 = parameters.index(filtered_list2[0])
        output_string2 = parameters[final_index2].split("=")[1]
    try:
        yuv_index = header_values.index('I|YUVFile')
        yuv_result_flag = 0
    except:
        yuv_result_flag = 1
    try:
        b_index = header_values.index('BitstreamFile')
        b_result_flag = 0
    except:
        b_result_flag = 1
    try:
        stream_md5_index = header_values.index('Stream MD5sum')
        stream_md5_flag = 0
    except:
        stream_md5_flag = 1
    try:
        hw_md5_index = header_values.index('HW MD5sum')
        hw_md5_flag = 0
    except:
        hw_md5_flag = 1
    try:
        index = header_values.index('Result')
        result_flag = 0
    except:
        result_flag = 1

    if yuv_result_flag != 1:
        yuv_result_cell = output_sheet.cell(row=cell.row,column=(yuv_index+1))
        yuv_result_cell.value = output_string2
    if b_result_flag != 1:
        b_result_cell = output_sheet.cell(row=cell.row,column=(b_index+1))
        b_result_cell.value = output_file
    if stream_md5_flag != 1:
        stream_md5_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
        stream_md5_cell.value = stream_md5_contents
    if hw_md5_flag != 1:
        hw_md5_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
        hw_md5_cell.value = hw_md5_content
    if omx_flag == 1:
        input_file_cell = output_sheet.cell(row=cell.row,column=(input_index_omx+1))
        input_file_cell.value = input_file_omx
        out_file_cell = output_sheet.cell(row=cell.row,column=(out_index_omx+1))
        out_file_cell.value = output_file_omx

    error_flag = parce_error(log_file, error_dict)
    if error_flag != 1 and result_flag != 1 and time_failure != 1 and md5_flag == 1:
        result_cell = output_sheet.cell(row=cell.row,column=(index+1))
        result_cell.value = "PASS"
        print("Result : ", result_cell.value)
    else:
        result_cell = output_sheet.cell(row=cell.row,column=(index+1))
        result_cell.value = "FAIL"
        print("Result : ", result_cell.value)
    output_workbook.save(output_xls)
    print("Completed----------------------", test_case, "-----------------------------------\n\n")

