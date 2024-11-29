import time
import os
import io
import sys
import openpyxl
import shutil
import fileinput
import subprocess
import time
import datetime
import argparse
import re
import glob
import signal

param_dict = {}
keys_values = {
        'GOP': ['Gop.DoubleRef', 'Gop.EnableLT', 'Gop.FreqIDR', 'Gop.FreqLT', 'Gop.FreqRP', 'Gop.GdrMode', 'Gop.Length',
        'Gop.NumB', 'Gop.TempDQP', 'Gop.WriteAVCHeaderSVC', 'GopCtrlMode'],

        'INPUT': ['CmdFile', 'I|CropHeight', 'I|CropPosX', 'I|CropPosY', 'I|CropWidth', 'I|Format', 'I|FrameRate', 'GMVFile', 'HDRFile',
        'I|Height', 'MapFile', 'QpTablesFolder', 'ROIFile', 'I|Width', 'I|YUVFile'],

        'DYNAMIC_INPUT': ['D|Height', 'D|Width', 'D|YUVFile'],

        'OUTPUT': ['BitstreamFile', 'O|CropHeight', 'O|CropPosX', 'O|CropPosY', 'O|CropWidth', 'O|Format', 'RecFile'],

        'RATE_CONTROL': ['BitRate', 'CPBSize', 'EnableSkip', 'FrameRate', 'InitialDelay', 'IPDelta', 'MaxBitRate',
        'MaxConsecutiveSkip', 'MaxPictureSize', 'MaxPictureSize.B', 'MaxPictureSize.I', 'MaxPictureSize.P',
        'MaxPictureSizeInBits', 'MaxPictureSizeInBits.B', 'MaxPictureSizeInBits.I', 'MaxPictureSizeInBits.P', 'MaxPSNR',
        'MaxQP', 'MaxQP.B', 'MaxQP.I', 'MaxQP.P', 'MinPSNR', 'MinQP', 'MinQP.B', 'MinQP.I',
        'MinQP.P', 'PBDelta', 'RateCtrlMode', 'ScnChgResilience', 'SCPrevention', 'SliceQP', 'UseGoldenRef'],

        'SETTINGS': ['AspectRatio', 'AvcLowLat', 'BitDepth', 'CabacInit', 'ChromaMode', 'ColourDescription', 'ColourMatrix',
        'Compression', 'ConstrainedIntraPred', 'CostMode', 'CuQpDeltaDepth',
        'DependentSlice', 'DirectMode', 'EnableAUD', 'EnableFillerData', 'EnableSEI', 'EntropyMode', 'FileScalingList', 'LambdaCtrlMode',
        'LambdaFactors', 'Level', 'LookAhead', 'LoopFilter', 'LoopFilter.BetaOffset', 'LoopFilter.CrossSlice', 'LoopFilter.CrossTile',
        'LoopFilter.TcOffset', 'NumCore', 'NumSlices', 'PCM', 'PicCbQpOffset', 'PicCrQpOffset', 'Profile', 'QPCtrlMode',
        'SAO', 'ScalingList', 'SCDFirstPass', 'SliceCbQpOffset', 'SliceCrQpOffset', 'SliceLat', 'SrcFormat', 'StartCode',
        'SubframeLatency', 'Tier', 'TransferCharac', 'TwoPass', 'TwoPassFile', 'UniformSliceType', 'UseL2C', 'VideoFullRange',
        'VideoMode', 'WeightedPred'],

        'RUN': ['BitrateFile', 'FirstPicture',
        'InputSleep', 'Loop', 'MaxPicture', 'ScnChgLookAhead', 'UseBoard']
}

# iterates through the key-value pairs, and it adds the entry to another dictionary `param_dict`.
for key, values in keys_values.items():
    param_dict[key] = values

# This dictionary contains predefined error messages,
# key : represents a general description of the error, and
# value : is a list of specific strings that are used to identify these errors in logs or output.
error_dict = {}
error_msg = {
        'Error in ctrlsw app': ['Error', 'error', 'ERROR'],
        'Assertion \'0\' failed': ['Assertion|failed', 'Assertion', 'assertion'],
        'Encoder Resource Error': ['Failed to create Encoder'],
        'CMA allocation Error': ['Cannot allocate memory'],
        'No QP File found': ['No QP File found'],
        'Unknown identifire please check property name': ['unknown identifier'],
        'Exception caught: No frame decoded': ['Exception caught: No frame decoded'],
        'I/p YUVFile not found': ['Exception caught: Can\'t open file for reading'],
        'Exception caught Error': ['Exception caught'],
        'Get higher Profile to support the usecase': ['getHevcMinimumProfile: Assertion \'0\' failed']
}

# iterates through the key-value pairs, and it adds the entry to another dictionary `error_dict`.
for key, values in error_msg.items():
    error_dict[key] = values

# Open an Excel file and Select a specific sheet from that file.
# Return both the selected sheet and the whole workbook for further manipulation.
def open_workbook(xls_file, xls_sheet):
    workbook = openpyxl.load_workbook(xls_file)
    workbook_sheet = workbook[xls_sheet]
    return workbook_sheet, workbook

# create a folder, checks if the folder already exists and It then returns the folder path for further use.
def extract_feature(sheet, row_no):
    cell = "A" + str(row_no)
    feature_cell = sheet[cell].value
    feature_cell = "Output/" + str(feature_cell.split(".")[1])
    try:
        os.mkdir(feature_cell)
        print("Created ", feature_cell, ": folder")
    except FileExistsError:
        print("folder", feature_cell, "already Exist do you want to continue and replace the data with new one? (y/n)")
        user_input2 = input()
        if user_input2.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()
    return feature_cell

# reads the header row of an Excel sheet, collects the values from each cell in that row, and returns them as a list.
def extract_header(sheet, header_row_number):
    param_values = []
    for cell in sheet[header_row_number]:
        cell_value = cell.value
        param_values.append(cell_value)
    return param_values

# function extracts parameter values from an Excel sheet, processes YUV file paths, and
# skips empty or invalid cells, ensuring valid inputs for further script execution.
# return processed list of parameter values
def extract_parameters_omx(sheet_omx, next_row_omx, cell_values_omx, output_folder_omx):
    #This function will be called in case of OpenMax TC only
    TC_No = sheet_omx.cell(row=next_row_omx, column=1).value
    a = 0
    i = 0
    j = 0
    skip = 0

    lines_omx = []

    for cell in sheet_omx[next_row_omx]:
        if cell_values_omx[i] == "Result" or cell_values_omx[i] == "TC_No" or cell_values_omx[i] == "HW MD5sum" or cell_values_omx[i] == "Comment":
            if cell.value == "PASS":
                skip = 1
                break
            i = i + 1
            continue

        #Non breaking space replace with None
        if cell.value == '\xa0':
            cell.value = None

        if cell.value == None and cell_values_omx[i] != "input_file" and cell_values_omx[i] != "out":
            i = i + 1
            continue

        if cell_values_omx[i] == "dma-out" or cell_values_omx[i] == "dma-in":
            if cell.value == 0:
                i = i + 1
                continue
            else:
                lines_omx.append("--" + str(cell_values_omx[i]))
                i = i + 1
                continue

        if cell_values_omx[i] == "Codec":
            Codec = cell.value
            value = "--" + str(cell.value)
            lines_omx.append(value)
            i = i + 1
            continue

        if cell_values_omx[i] == "input_file":
            encoded_omx_file = "/mnt/build/ssw_vcu/yashl/VCU2/VNC_testing/IPP_platform_logs/omx_support/Output/OpenMax"
            if 'hevc' in str(Codec):
                input_omx_file = str(encoded_omx_file) + "/" + str(TC_No) + "/" + str(TC_No) + ".hevc"
            elif 'avc' in str(Codec):
                input_omx_file = str(encoded_omx_file) + "/" + str(TC_No) + "/" + str(TC_No) + ".avc"
            else:
                input_omx_file = "/mnt/build/ssw_vcu/yashl/VCU2/VNC_testing/IPP_platform_logs/omx_support/Output/OpenMax/TC_066/TC_066.avc"
            value = str(input_omx_file)
            lines_omx.append(value)
            i = i + 1
            continue

        if cell_values_omx[i] == "out":
            out_file = str(log_folder) + "/" + str(TC_No) + "/" + str(TC_No) + ".yuv"
            value = "--" + str(cell_values_omx[i]) + " " + str(out_file)
            lines_omx.append(value)
            i = i + 1
            continue

        i = i + 1
        continue

    try:
        os.mkdir(str(output_folder_omx) + "/" + str(TC_No))
        print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
    except FileExistsError:
        print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
        user_input = input()
        if user_input.lower() == 'y':
            pass
        else:
            print("Program closing")

            exit()
    return lines_omx

def extract_parameters(sheet, next_row, cell_values, output_folder):

    #In XLS we will be using row 4 as the heading of the parameters
    TC_No = sheet.cell(row=next_row, column=1).value
    source_file = 'input_files/input.cfg'
    a = 0
    i = 0
    j = 0

    #lines will store the lines made from parsing the table and that we will insert in the table
    lines = []

    #target_text will be holding the cfg section info that in which section line will be apended.
    target_text = []
    avc_flag = 0

    #This loop generates the lines that needs to be updated on cfg file
    for cell in sheet[next_row]:
        value = str(cell_values[i]) + "      =      " + str(cell.value) + " "
        lines.append(value)
        i = i+1

    #This block of code generates the cfg files for each testcase
    try:
        os.mkdir(str(output_folder) + "/" + str(TC_No))
        print("Created ", str(TC_No), "folder: Head over to this folder for more TC related information and output files")
    except FileExistsError:
        print("Output folder ", str(TC_No), ": alreaady Exist do you want to continue and replace the data with new one? (y/n)")
        user_input = input()
        if user_input.lower() == 'y':
            pass
        else:
            print("Program closing")
            exit()

    #This block compares the target_text with sections inside cfg and gets the lines no. and it will append the line on next line of
    #matching section

    return lines

# Identify known issues or failures by checking logs predefined error message
def parce_error(file_path, error_dict):
    with open(file_path, 'r') as file:
        file_contents = file.read()
        for error_message, error_keywords in error_dict.items():
            for keyword in error_keywords:
                if keyword in file_contents:
                    print(f"Error: {error_message}")
                    error = 1
                    return error



#----------------------------------------------------------------------------------------------------------------------#

parser = argparse.ArgumentParser(description='Testcase_automation_V1.0', add_help=True)

parser.add_argument('-f', '--file', help='Specify a input XLS file for testcases')
parser.add_argument('-s', '--sheet', help='Specify a input XLS file\'s sheet for testcases')
parser.add_argument('-o', '--output', action='store_true', default=False, help='Select this option to store the output file')
parser.add_argument('-tc', '--tc_no', help='Specify individual TC_No that u want to run e.g -tc TC_0001')

args = parser.parse_args()

file_option = args.file
sheet_option = args.sheet
output_option = args.output

try:
    os.mkdir("Output")
    print("Created Output folder: Head over to this folder for more TC related information and output files")
except FileExistsError:
    user_input1 = input("Output folder alreaady Exist do you want to continue and replace the data with new one? (y/n)")
    if user_input1.lower() == 'y':
        pass
    else:
        print("Program closing")
        exit()

CWD = os.getcwd()
omx_flag = 0
hw_md5_flag = 0
orignal_xls = args.file
output_xls = "Output/output.xlsx"

#Path of Argon bitstream files
folder_path_argon = "/mnt/group/siv3/staff/andreis/vcu_deploy/deliveries/database/input/decoder/complex_dec"
#Path of Fuzz and currupted files path
folder_path_fuzz = "/mnt/group/siv3/staff/andreis/vcu_deploy/deliveries/database/input/fuzz_all"
#Path of Conformance HEVC
folder_path_conf = "/mnt/group/siv3/staff/andreis/vcu_deploy/deliveries/database/md5sum/xilinx_private/ctrlSW/stream_level/ctrlSW/input/decoder"
#Path of Driver level
folder_path_driver_level = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Color_Format/Output/Color_Format"
folder_path_driver_level_yuv = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Decoder/Driver_level_0818/Output/Decoder_Driver_Level"
#Path of encoded file for Latency Mode
folder_path_low_latency = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Encoder/Parameters/Output"
#Path of yuv file for Latency Mode
folder_path_latency_mode_yuv = "/mnt/everest/ssw_multimedia_bkup/VCU2/regression_logs/Decoder/Dec_Low_latency/Output"

#Coping orignal xls into output_xls so we can modify content or put result in output xls
shutil.copy2(orignal_xls, output_xls)

output_sheet, output_workbook = open_workbook(str(output_xls), str(args.sheet))
#It's bit depth flag if no Bitdepth parametre in xls it will remain zero
bd_flag = 0

sheet, new_workbook = open_workbook(str(args.file), str(args.sheet))
#next_row, header_values, Heading_cell = extract_headers(sheet)
#We are iterating over every row for column A
for cell in sheet['A']:
    if sheet_option == "Dec_Conformance":
        no_stream_md5_file = 0
    if sheet_option == "OpenMax":
        omx_flag = 1

    fill_color = cell.fill.start_color.rgb

   #We are checking for red color cell if red color cell found we are done with all the testcases and program will close
    if sheet[cell.coordinate].fill.start_color.rgb == 'FFFF0000':
        print("\n\n -/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/-/- \n\n")
        break

    #If there is cell with no black color and it's value is None we skip that row
    if sheet[cell.coordinate].fill.start_color.rgb != 'FF000000' and cell.value is None:
        continue

    #If we find Black color cell we will have feature cell in next row
    if sheet[cell.coordinate].fill.start_color.rgb == 'FF000000':
        #Enabling flag so in next row we will extract feature name
        extract_feature_flag = 1
        continue

    if extract_feature_flag == 1:
        feature_folder = extract_feature(sheet, cell.row)
        log_folder = str(CWD) + "/" + str(feature_folder)
        extract_feature_flag = 0
        #We got the feature name enabling this flag as in next row we will extract the headers of testcase
        extract_header_flag = 1
        if feature_folder == "Output/Low_latency":
            latency_flag = 1
        else:
            latency_flag = 0
        continue

    if extract_header_flag == 1:
        header_values = extract_header(sheet, cell.row)
        extract_header_flag = 0
        continue

    if args.tc_no is not None:
        if cell.value != args.tc_no:
            continue

#--------------------------------------------------------------------------------------------------------------------    
    if omx_flag == 1:
         parameters = extract_parameters_omx(sheet, cell.row, header_values, log_folder)
    else :
        parameters = extract_parameters(sheet, cell.row, header_values, log_folder)
#--------------------------------------------------------------------------------------------------------------------    
    if omx_flag == 1:
        substring = "out"
        filtered_list = [element for element in parameters if substring in element]
        if filtered_list:
            final_index_omx = parameters.index(filtered_list[0])
            output_string_omx = parameters[final_index_omx].split(" ")[1]
            output_file_omx = output_string_omx.split("/")[-1]

        input_file_param_index = [i for i, j in enumerate(parameters) if '.hevc' in j or '.avc' in j]
        input_file_param_index = input_file_param_index[0]

        if 'input_file' in header_values:
            input_index_omx = header_values.index('input_file')
            input_file_omx = parameters[input_file_param_index]
        if 'out' in header_values:
            out_index_omx = header_values.index('out')

    substring = "Bitstream"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
        bitstream_file = parameters[final_index].split("=")[1]
        bitstream_file = bitstream_file.replace(" ","")

    substring = "TC_No"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        final_index = parameters.index(filtered_list[0])
        testcase_no = parameters[final_index].split("=")[1]
        testcase_no = testcase_no.replace(" ","")

    if sheet_option == "Dec_Driver_level" or sheet_option == "Latency_mode":
        substring = "yuv file"
        filtered_list = [element for element in parameters if substring in element]
        if filtered_list:
            yuv_file_index = parameters.index(filtered_list[0])
            yuv_file = parameters[yuv_file_index].split("=")[1]
            yuv_file = yuv_file.replace(" ","")
    
    substring = "HW MD5sum"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        hw_md5_index = parameters.index(filtered_list[0])

    substring = "Argon Md5Sum"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        md5_index = parameters.index(filtered_list[0])
        argon_md5sum = parameters[md5_index].split("=")[1]
        argon_md5sum = argon_md5sum.replace(" ","")

    substring = "Category"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        cat_index = parameters.index(filtered_list[0])
        cat_folder = parameters[cat_index].split("=")[1]
        cat_folder = cat_folder.replace(" ","")
        bd = cat_folder.split("_")[0]
        if bd == "main10":
            bd_flag = 1

    substring = "Codec"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        codec_index = parameters.index(filtered_list[0])
        codec_value = parameters[codec_index].split("=")[1]
        codec_value = codec_value.replace(" ","")

    substring = "BitDepth"
    filtered_list = [element for element in parameters if substring in element]
    if filtered_list:
        bitdepth_index = parameters.index(filtered_list[0])
        bitdepth_value = parameters[bitdepth_index].split("=")[1]
        bitdepth_value = bitdepth_value.replace(" ","")

    if sheet_option == "Dec_Conformance":
        for root, dirs, files in os.walk(folder_path_conf):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)
                stream_md5_file = file_path.rsplit('.', 1)[0] + '.md5'
                if os.path.exists(stream_md5_file):
                    pass
                else:
                    print("File not found : ", stream_md5_file)
                    no_stream_md5_file = 1
                try:
                    with open(stream_md5_file, 'r', encoding='utf-8') as file:
                        stream_md5_contents = file.read()
                        stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
                except FileNotFoundError:
                    print("File not found")
                except IOError:
                    print("Error reading file")
        

    if sheet_option == "Dec_Argon":
        final_bitstream_path = folder_path_argon + "/" + cat_folder
        for root, dirs, files in os.walk(final_bitstream_path):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)

    if sheet_option == "Dec_Driver_level":
        if codec_value == 'MJPEG':
            final_bitstream_path = '/mnt/everest/ssw_multimedia_bkup/VCU2/MJPEG_StreamFile'
        elif codec_value == 'JPG':
            final_bitstream_path = '/mnt/group/siv3/staff/andreis/sibridge/yashl/VCU2/jpegs'
        else:
            final_bitstream_path = folder_path_driver_level + "/" + testcase_no
        for root, dirs, files in os.walk(final_bitstream_path):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)
        md5_file = testcase_no + ".md5"
        folder_path_driver_level_yuv_new = folder_path_driver_level_yuv + "/" + testcase_no
        for root, dirs, files in os.walk(folder_path_driver_level_yuv_new):
            if md5_file in files:
                md5_file_path = os.path.join(root, testcase_no + ".md5")
                stream_md5_file = md5_file_path.rsplit('.', 1)[0] + '.md5'
                if os.path.exists(stream_md5_file):
                    pass
                else:
                    print("File not found : ", stream_md5_file)
                    no_stream_md5_file = 1
                try:
                    with open(stream_md5_file, 'r', encoding='utf-8') as file:
                        stream_md5_contents = file.read()
                        stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
                        stream_md5_contents = stream_md5_contents.split(" ")[0]
                except IOError:
                    print("Error reading file")


    if sheet_option == "Latency_mode":
        if latency_flag:
            final_bitstream_path = folder_path_low_latency + "/RateControl"
            folder_path_latency_mode_yuv_new = folder_path_latency_mode_yuv + "/Low_latency"
        else:
            final_bitstream_path = folder_path_low_latency + "/GOP"
            folder_path_latency_mode_yuv_new = folder_path_latency_mode_yuv + "/Reduced_latency"
        for root, dirs, files in os.walk(final_bitstream_path):
            if bitstream_file in files:
                file_path = os.path.join(root, bitstream_file)
                break
            else:
                file_path = None
        md5_file = testcase_no + ".md5"
        folder_path_latency_mode_yuv_new = folder_path_latency_mode_yuv_new + "/" + testcase_no
        for root, dirs, files in os.walk(folder_path_latency_mode_yuv_new):
            if md5_file in files:
                md5_file_path = os.path.join(root, testcase_no + ".md5")
                stream_md5_file = md5_file_path.rsplit('.', 1)[0] + '.md5'
                if os.path.exists(stream_md5_file):
                    pass
                else:
                    print("File not found : ", stream_md5_file)
                    no_stream_md5_file = 1
                try:
                    with open(stream_md5_file, 'r', encoding='utf-8') as file:
                        stream_md5_contents = file.read()
                        stream_md5_contents = re.sub(r'\s+', ' ', stream_md5_contents).strip()
                        stream_md5_contents = stream_md5_contents.split(" ")[0]
                except IOError:
                    print("Error reading file")

    if sheet_option == "Dec_Fuzz":
        for root, dirs, files in os.walk(folder_path_fuzz):
            if bitstream_file in dirs:
                fuzz_file_path = os.path.join(root, bitstream_file)
                fuzz_file_list = os.listdir(fuzz_file_path)
                if len(fuzz_file_list) != 0:
                    fuzz_file_name = fuzz_file_list[0]
                    file_path = os.path.join(fuzz_file_path, fuzz_file_name)

    if omx_flag == 1:
        if input_file_omx is None:
            print("Input file not found")
    else:
        if file_path is None:
            print("Bistream file not found")

    test_case = cell.value
    print("Running----------------------", test_case, "-----------------------------------\n\n")

    log_file = log_folder + "/" + cell.value + "/" + str(cell.value) + ".txt"
    log_file = log_file.replace(" ","")
    md5sum_file = log_file.split(".")[0] + ".md5"

    with open(log_file, "w") as file:
        current_time = datetime.datetime.now()
        mem_command = "cat /proc/meminfo"
        process = subprocess.Popen(mem_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        process.wait()
        #this the maximum time we will wait for 1 usecase
        deadline = current_time + datetime.timedelta(minutes=5)
        if sheet_option == "Dec_Conformance":
            if bitdepth_value == "10":
                if codec_value == "AVC":
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
                else:
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
            else:
                if codec_value == "AVC":
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file
                else:
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file

        if sheet_option == "Dec_Argon":
            if bd_flag == 1:
                command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
            else:
                command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file

        if sheet_option == "Dec_Fuzz":
            command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -noyuv --md5 " + md5sum_file

        if sheet_option == "Dec_Driver_level" or sheet_option == "Latency_mode":
           # if codec_value != 'MJPEG':
           #     md5_command = ['md5sum', yuv_file]
           #     print(md5_command)
           #     md5_process = subprocess.Popen(md5_command, stdout=subprocess.PIPE)
           #     output, error = md5_process.communicate()
           #     if md5_process.returncode == 0:
           #         md5sum_output = output.decode().split()[0]
     #     #          print(f'MD5sum: {md5sum_output}')
           #     else:
           #         print(f'Error: {error.decode()}')
            if bitdepth_value == "10":
                if codec_value == "AVC":
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
                elif codec_value == "MJPEG" or codec_value == "JPG":
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -jpeg -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
                else:
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 10" + " -noyuv --md5 " + md5sum_file
            elif bitdepth_value == "12":
                if codec_value == "AVC":
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 12" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 12" + " -noyuv --md5 " + md5sum_file
                elif codec_value == "MJPEG" or codec_value == "JPG":
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -jpeg -in " + str(file_path) + " -bd 12" + " -noyuv --md5 " + md5sum_file
                else:
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 12" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 12" + " -noyuv --md5 " + md5sum_file
            else:
                if codec_value == "AVC":
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -avc -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file
                elif codec_value == "MJPEG" or codec_value == "JPG":
                    command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -jpeg -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file
                else:
                    if latency_flag:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file + " -lowlat"
                    else:
                        command = "ctrlsw_decoder --embedded --device /dev/al_d3xx -in " + str(file_path) + " -bd 8" + " -noyuv --md5 " + md5sum_file

        if sheet_option == 'Performance':
            for param in parameters:
                if 'TC_No' in param:
                    tc_no = param.split('=')[-1].strip()

                if 'BitstreamFile' in param:
                    encoded_file = param.split('=')[-1].strip()
                    if encoded_file.endswith('.avc'):
                        omx_dec = 'omxh264dec'
                        h26_parse = 'h264parse'
                    elif encoded_file.endswith('.hevc'):
                        omx_dec = 'omxh265dec'
                        h26_parse = 'h265parse'

            source_path = "/mnt/build/ssw_vcu/yashl/VCU2/VNC_testing/IPP_platform_logs/Performance/Output/Performance_filesrc"
            command = f'gst-launch-1.0 filesrc location={source_path}/{tc_no}/{encoded_file} ! {h26_parse} ! {omx_dec} internal-entropy-buffers=5 !  video/x-raw ! queue max-size-bytes=0 ! fpsdisplaysink name=fpssink text-overlay=false video-sink="fakesink" sync=false -v'

        if omx_flag == 1:
            command = "omx_decoder " + str(parameters[input_file_param_index])

            for x in parameters:
                if '.hevc' in x:
                    continue
                elif '.avc' in x:
                    continue
                elif 'out' in x:
                    bitstream_omx = str(x)
                    bitstream_omx = bitstream_omx.split(" ")[1]
                    command = command + " " + str(x)
                else:
                    command = command + " " + str(x)
        print(command,"\n\n")
        process = subprocess.Popen(command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        pid = process.pid
        #Polling here until the encoding or decoding is Done
        while process.poll() is None:
            time.sleep(5)
#            if datetime.datetime.now() > deadline:
#                time_failure = 1
#                break
        process = subprocess.Popen(mem_command, shell=True, stdout=file, stderr=subprocess.STDOUT, text=True)
        process.wait()

        if omx_flag == 1:
            bitstream_file = bitstream_omx

            md5_command = ['md5sum', bitstream_file]
            md5_process = subprocess.Popen(md5_command, stdout=subprocess.PIPE)
            output, error = md5_process.communicate()
            if md5_process.returncode == 0:
                hw_md5_flag = 1
                hw_md5_content = output.decode().split()[0]
            else:
                print(f'md5 Error: {md5_command}')
                hw_md5_content = " "
            with open (md5sum_file, 'w') as file:
                file.writelines(hw_md5_content)
            if os.path.exists(bitstream_file):
                os.remove(bitstream_file)
    try:
        with open(md5sum_file, 'r', encoding='utf-8') as file:
            hw_md5_contents = file.read()
            hw_md5_contents = hw_md5_contents.strip()
    except FileNotFoundError:
        print("File not found")

    if sheet_option == "Dec_Conformance" or sheet_option == "Dec_Driver_level" or sheet_option == "Latency_mode" :
        try:
            stream_md5_index = header_values.index('Stream MD5sum')
            stream_md5_flag = 0
        except:
            stream_md5_flag = 1

    try:
        index = header_values.index('Result')
        result_flag = 0
    except:
        result_flag = 1

    try:
        hw_md5_index = header_values.index('HW MD5sum')
        hw_md5_result_flag = 0
    except:
        hw_md5_result_flag = 1

    if sheet_option == "Dec_Argon":
        if argon_md5sum == hw_md5_contents:
            md5_flag = 1
        else:
            md5_flag = 0

    if sheet_option == "Dec_Driver_level"  or sheet_option == "Latency_mode" :
        if stream_md5_contents == hw_md5_contents:
            md5_flag = 1
        else:
            md5_flag = 0

    if sheet_option == "Dec_Conformance":
        if stream_md5_contents == hw_md5_contents:
            matching_md5_flag = 1
        else:
            matching_md5_flag = 0

    error_flag = parce_error(log_file, error_dict)

    if sheet_option == "Dec_Conformance":
        if error_flag != 1 and result_flag != 1 and matching_md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
            result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
            result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                result_cell.value = hw_md5_contents
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"
            if no_stream_md5_file != 1:
                result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
                result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                 result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                 result_cell.value = hw_md5_contents

    if sheet_option == "Dec_Argon":
        if error_flag != 1 and result_flag != 1 and md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"

    if sheet_option == "Dec_Driver_level" or sheet_option == "Latency_mode":
        if error_flag != 1 and result_flag != 1 and md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
            result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
            result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                result_cell.value = hw_md5_contents
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"
            result_cell = output_sheet.cell(row=cell.row,column=(stream_md5_index+1))
            result_cell.value = stream_md5_contents
            if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
                result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
                result_cell.value = hw_md5_contents

    if sheet_option == "Dec_Fuzz":
        if error_flag != 1 and result_flag != 1:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"

    if sheet_option == "Dec_Argon":
        if hw_md5_result_flag != 1 and hw_md5_contents != "md5sum:":
            result_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
            result_cell.value = hw_md5_contents

    if omx_flag == 1:
        if hw_md5_flag == 1:
            hw_md5_cell = output_sheet.cell(row=cell.row,column=(hw_md5_index+1))
            hw_md5_cell.value = hw_md5_content
        input_file_cell = output_sheet.cell(row=cell.row,column=(input_index_omx+1))
        input_file_cell.value = input_file_omx
        out_file_cell = output_sheet.cell(row=cell.row,column=(out_index_omx+1))
        out_file_cell.value = output_file_omx
        if error_flag != 1 and result_flag != 1 and hw_md5_flag != 0:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "PASS"
            print("Result : ", result_cell.value)
        else:
            result_cell = output_sheet.cell(row=cell.row,column=(index+1))
            result_cell.value = "FAIL"
            print("Result : ", result_cell.value)

    output_workbook.save(output_xls)

    print("Completed----------------------", test_case, "-----------------------------------\n\n")


